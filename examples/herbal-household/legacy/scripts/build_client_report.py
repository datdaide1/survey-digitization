#!/usr/bin/env python3
"""Task 6 — Báo cáo thống kê gửi khách (XLSX + DOCX).

Đọc output/combined.csv (chạy scripts/build_stats_layer.py trước) + output/full/*.json
(cho sheet "Dữ liệu đã số hóa", có PII theo §10) -> reports/bao-cao-khao-sat.xlsx + .docx.

Đủ 7 tầng theo docs/implement-plan-statistics-and-client-report.md: Tầng 1 (tần suất +
mô tả, sống bằng công thức Excel), Tầng 3 (ma trận liên quan), Tầng 4-5 (effect size + so
sánh nhóm phi tham số, gộp 1 sheet), Tầng 6 (Cronbach's alpha + chỉ số tổng hợp), Tầng 7
(factor analysis + cluster analysis, tĩnh — xem scripts/lib/report/advanced_sheet.py).

26/07: bỏ Tầng 2 (cross-tab theo tỉnh/xã) + radar Q32 so theo tỉnh — quyết định của user,
không cần so sánh giữa các vùng nữa (xem memory/[[project-survey-no-region-comparison]]).
`write_crosstab_sheet`/`write_q32_radar` vẫn còn trong xlsx_writer.py (không xoá code),
chỉ không gọi ở đây nữa — dễ bật lại nếu sau này cần.

Chạy:
  & "E:\\anaconda3\\envs\\survey-digitizer\\python.exe" scripts/build_stats_layer.py
  & "E:\\anaconda3\\envs\\survey-digitizer\\python.exe" scripts/build_client_report.py
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import pandas as pd  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from lib.flatten import flatten_full  # noqa: E402
from lib.records import iter_full_records  # noqa: E402
from lib.report.advanced_sheet import write_advanced_sheet  # noqa: E402
from lib.report.association_sheet import write_association_sheet  # noqa: E402
from lib.report.docx_writer import build_docx  # noqa: E402
from lib.report.effect_size_sheet import write_effect_size_sheet  # noqa: E402
from lib.report.reliability_sheet import write_reliability_sheet  # noqa: E402
from lib.report.xlsx_writer import (  # noqa: E402
    write_chart_sheet,
    write_data_sheet,
    write_frequency_sheet,
    write_full_data_sheet,
)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

CHART_SPECS = [
    {"column": "Q2_tuoi", "kind": "bar"},  # bảng mô tả, không có block tần suất -> bỏ qua ở write_chart_sheet
    {"column": "age_bracket", "kind": "bar"},
    {"column": "Q4", "kind": "bar"},
    {"column": "education_grade_bracket", "kind": "bar"},
    {"column": "Q6", "kind": "pie"},
    {"column": "marriage_age_bracket", "kind": "bar"},
    {"column": "Q7_cay_duoc_lieu", "kind": "bar"},
    {"column": "Q8", "kind": "bar"},
    {"column": "experience_years_bracket", "kind": "pie"},
    {"column": "Q10", "kind": "bar"},
    {"column": "Q11_hoi_phu_nu", "kind": "bar"},
]


def build_workbook(combined_csv: str, full_dir: str) -> tuple[Workbook, dict, dict, dict, dict]:
    df = pd.read_csv(combined_csv)
    column_order = [c for c in df.columns if c != "record_id"]

    full_rows = [flatten_full(r) for r in iter_full_records(full_dir)]
    full_df = pd.DataFrame(full_rows)

    wb = Workbook()
    wb.remove(wb.active)  # bỏ sheet mặc định trống

    write_full_data_sheet(wb, full_df)
    data_refs = write_data_sheet(wb, "Dữ liệu (ẩn danh)", df, hidden=True)
    locations = write_frequency_sheet(wb, data_refs, df, column_order)
    write_chart_sheet(wb, locations, CHART_SPECS)
    association_info = write_association_sheet(wb, df, column_order)
    effect_size_info = write_effect_size_sheet(wb, df)
    reliability_info = write_reliability_sheet(wb, df, data_refs)
    advanced_info = write_advanced_sheet(wb, df)

    wb.active = wb.sheetnames.index("Thống kê tổng hợp")
    return wb, association_info, effect_size_info, reliability_info, advanced_info


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--combined-csv", default="output/combined.csv")
    ap.add_argument("--full-dir", default="output/full")
    ap.add_argument("--xlsx-out", default="reports/bao-cao-khao-sat.xlsx")
    ap.add_argument("--docx-out", default="reports/bao-cao-khao-sat.docx")
    args = ap.parse_args()

    wb, association_info, effect_size_info, reliability_info, advanced_info = build_workbook(args.combined_csv, args.full_dir)
    xlsx_path = Path(args.xlsx_out)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"OK: đã ghi {xlsx_path} — sheet: {', '.join(wb.sheetnames)}")

    df = pd.read_csv(args.combined_csv)
    docx_path = Path(args.docx_out)
    build_docx(
        df, docx_path,
        association_info=association_info,
        effect_size_info=effect_size_info,
        reliability_info=reliability_info,
        advanced_info=advanced_info,
    )
    print(f"OK: đã ghi {docx_path}")
    sys.exit(0)


if __name__ == "__main__":
    main()
