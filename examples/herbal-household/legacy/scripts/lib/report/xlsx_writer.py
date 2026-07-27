"""Dựng workbook Excel gửi khách — sheet dữ liệu + Tầng 1 (tần suất/mô tả) + Tầng 2
(cross-tab theo tỉnh) bằng CÔNG THỨC EXCEL GỐC (COUNTIF/COUNTIFS/AVERAGE/...), sống
100% theo đúng quyết định §2/§4.1 của docs/implement-plan-statistics-and-client-report.md:
mọi con số trong sheet thống kê là formula tham chiếu tới sheet dữ liệu ẩn danh, không
phải giá trị Python dán tĩnh — khách sửa 1 ô dữ liệu, số thống kê tự cập nhật.

Sheet "Dữ liệu đã số hóa" (có PII, §10) KHÔNG sống theo nghĩa đó — đó là bản dump đầy
đủ để chứng minh đã số hóa từ phiếu thật, không phải nguồn cho công thức.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, RadarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..flatten import DEVICE_GRID_FIELDS, MATRIX_ROW_FIELDS, MULTI_SELECT_FIELDS
from .codebook import build_codebook
from .crosstab import PROVINCE_LABELS
from .frequency import CONTINUOUS_COLUMNS, category_rows

CODEBOOK = build_codebook()

# 26/07 (tối, phản hồi khách — "làm gì có nhà phân tích thống kê nào" trình bày multi-select
# bằng cách tách riêng "Có chọn/Không chọn" cho TỪNG lựa chọn): các cột đã nổ từ 1 câu
# multi-select/ma trận/device-grid PHẢI gộp lại thành 1 bảng/1 khối duy nhất (mỗi lựa chọn/
# dòng là 1 hàng trong CÙNG bảng đó), không phải N khối rời rạc mỗi khối 1 lựa chọn — xem
# _write_multiselect_freq_block/_write_matrix_freq_block bên dưới. Set tra cứu nhanh cột nào
# thuộc nhóm nào, dùng trong write_frequency_sheet để bỏ qua các cột này khi lặp column_order
# rồi viết gộp riêng.
GROUPED_QUESTION_COLUMNS: dict[str, str] = {}  # column_name -> qid gốc
for _qid, _opts in MULTI_SELECT_FIELDS.items():
    for _code in _opts:
        GROUPED_QUESTION_COLUMNS[f"{_qid}_{_code}"] = _qid
for _qid, _rows in MATRIX_ROW_FIELDS.items():
    for _code in _rows:
        GROUPED_QUESTION_COLUMNS[f"{_qid}_{_code}"] = _qid
for _qid, _spec in DEVICE_GRID_FIELDS.items():
    for _row in _spec["rows"]:
        for _col in _spec["columns"]:
            GROUPED_QUESTION_COLUMNS[f"{_qid}_{_row}_{_col}"] = _qid
    if _spec.get("extra_option"):
        GROUPED_QUESTION_COLUMNS[f"{_qid}_{_spec['extra_option']}"] = _qid

# Nhãn gốc (không có hậu tố "— lựa chọn cụ thể") cho mỗi qid đã gộp — dùng làm tiêu đề
# chart, vì CODEBOOK không có entry cho bare "Q7"/"Q28"/... (chỉ có entry cho từng cột đã nổ).
GROUP_BASE_LABELS: dict[str, str] = {}
for _qid, _opts in MULTI_SELECT_FIELDS.items():
    GROUP_BASE_LABELS[_qid] = CODEBOOK[f"{_qid}_{_opts[0]}"]["label"].split(" — ", 1)[0]
for _qid, _rows in MATRIX_ROW_FIELDS.items():
    GROUP_BASE_LABELS[_qid] = CODEBOOK[f"{_qid}_{_rows[0]}"]["label"].split(" — ", 1)[0]
for _qid, _spec in DEVICE_GRID_FIELDS.items():
    GROUP_BASE_LABELS[_qid] = CODEBOOK[f"{_qid}_{_spec['rows'][0]}_{_spec['columns'][0]}"]["label"].split(" — ", 1)[0]

# 26/07 (tối, phản hồi khách "làm đẹp/chuyên nghiệp hơn"): font Arial xuyên suốt thay vì để
# openpyxl mặc định Calibri khi Font() không khai báo name — áp dụng ở đây (nguồn dùng chung
# của mọi sheet trụ cột, xem pillar_xlsx.py import lại các hằng số này).
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13, color="1F4E78")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="666666")
DEFAULT_FONT = Font(name=FONT_NAME, size=10)


def _excel_str(value: Any) -> str:
    """Escape " thành "" cho literal chuỗi trong công thức Excel."""
    return str(value).replace('"', '""')


def _fixed_rows_def(meta: dict[str, Any], column: str, df: pd.DataFrame) -> list[tuple[Any, str]]:
    """Danh sách (code, label) theo đúng "kind" của biến — KHÔNG dựa vào isinstance()
    trên giá trị đọc từ pandas (numpy.int64 không phải subclass của int, dễ lệch nhánh
    một cách âm thầm) — 2 kind có giá trị cố định (binary/boolean) dùng literal Python
    thường; chỉ "categorical" mới cần soi dữ liệu thật qua category_rows()."""
    if meta["kind"] == "boolean":
        return [(True, "Có"), (False, "Không")]
    if meta["kind"] == "binary":
        return [(1, "Có chọn"), (0, "Không chọn")]
    return category_rows(column, df[column])


def _criteria_literal(meta: dict[str, Any], code: Any) -> str:
    """Literal dùng trong COUNTIF/COUNTIFS — số cho binary (không quote, tránh phụ
    thuộc coi-chuỗi-là-số ngầm định của Excel), TRUE/FALSE cho boolean, chuỗi có
    escape cho categorical.

    26/07 (phát hiện khi build sheet trụ cột mới — BUG có sẵn từ trước, ảnh hưởng cả
    age_bracket/marriage_age_bracket/experience_years_bracket ở "Thống kê tổng hợp"):
    Excel/LibreOffice diễn giải criteria bắt đầu bằng <, >, = như TOÁN TỬ SO SÁNH, không
    phải so khớp văn bản — COUNTIF(range,"<35") thực chất hỏi "giá trị < 35" (luôn sai/0
    khi range là text), KHÔNG PHẢI "bằng chữ '<35'". Test tay bằng recalc.py xác nhận:
    COUNTIF(A1:A7,"<35") = 0 trong khi COUNTIF(A1:A7,"=<35") = 2 (đúng). Cách sửa chuẩn
    của Excel: thêm dấu "=" ở đầu literal để ép so khớp CHÍNH XÁC văn bản, an toàn cho
    MỌI chuỗi kể cả chuỗi không có ký tự đặc biệt (đã kiểm tra "=foo" cho kết quả giống
    hệt "foo") — nên áp dụng "=" cho MỌI literal categorical, không chỉ case có vấn đề."""
    if meta["kind"] == "boolean":
        return "TRUE" if code is True else "FALSE"
    if meta["kind"] == "binary":
        return str(code)
    return f'"={_excel_str(code)}"'


def _combo_criteria_literals(meta: dict[str, Any], code: Any) -> list[str]:
    """Literal(s) cần CỘNG DỒN để đếm đúng 1 lựa chọn categorical, kể cả khi phiếu tick
    ≥2 ô trên câu chỉ cho chọn 1 và giá trị được ghi dạng 'code1+code2' (xem
    scripts/lib/records.as_single_category, và scripts/lib/report/frequency.py — bản
    Python đã tính cùng nguyên tắc: 1 phiếu combo được tính vào MỌI lựa chọn nó tick,
    không tạo hạng mục tổ hợp riêng). Ranh giới bằng dấu '+' để khớp CHÍNH XÁC 1 code,
    tránh COUNTIF khớp nhầm theo kiểu substring (vd 'vo' không khớp bên trong 'khong_vo'
    vì không có dấu '+' bao quanh). binary/boolean không thể có '+' nên chỉ 1 literal."""
    if meta["kind"] != "categorical":
        return [_criteria_literal(meta, code)]
    # 26/07: "=" ở đầu mỗi literal (kể cả bản có wildcard "*") — cùng lý do đã ghi ở
    # _criteria_literal, đã kiểm tra tay bằng recalc.py: "=<35+*" vẫn khớp đúng
    # "<35+abc" (wildcard "*" vẫn hoạt động bình thường cùng dấu "=" ép so khớp chính
    # xác phần chữ), không làm hỏng cách khớp tổ hợp hiện có.
    base = _excel_str(code)
    return [f'"={base}"', f'"={base}+*"', f'"=*+{base}"', f'"=*+{base}+*"']


def _count_formula(meta: dict[str, Any], data_range: str, code: Any, extra_criteria: str | None = None) -> str:
    """COUNTIF/COUNTIFS cộng dồn qua mọi literal ở _combo_criteria_literals — thay cho
    1 COUNTIF/COUNTIFS đơn lẻ trước đây (bỏ sót phiếu multi-mark vì so khớp CHÍNH XÁC
    chuỗi, 'nong_dan' không khớp ô có giá trị 'nong_dan+buon_ban')."""
    parts = []
    for literal in _combo_criteria_literals(meta, code):
        if extra_criteria:
            parts.append(f'COUNTIFS({data_range},{literal},{extra_criteria})')
        else:
            parts.append(f'COUNTIF({data_range},{literal})')
    return "+".join(parts)


@dataclass
class DataRef:
    sheet: str
    col_letter: str
    first_row: int
    last_row: int

    @property
    def range(self) -> str:
        return f"'{self.sheet}'!${self.col_letter}${self.first_row}:${self.col_letter}${self.last_row}"

    def eq(self, literal_formula_or_str) -> str:
        """range == giá trị — dùng trong COUNTIF điều kiện."""
        return self.range


def write_data_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, hidden: bool = False) -> dict[str, DataRef]:
    ws = wb.create_sheet(sheet_name)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col).font = HEADER_FONT
        ws.cell(row=1, column=j).fill = HEADER_FILL
    for i, (_, record) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            value = record[col]
            if pd.isna(value):
                continue  # để trống hẳn -> COUNTA/COUNTBLANK hoạt động đúng
            if isinstance(value, (bool,)):
                ws.cell(row=i, column=j, value=bool(value))
            elif isinstance(value, (int,)):
                ws.cell(row=i, column=j, value=int(value))
            elif isinstance(value, float):
                ws.cell(row=i, column=j, value=float(value))
            else:
                ws.cell(row=i, column=j, value=str(value))
    ws.freeze_panes = "A2"
    if hidden:
        ws.sheet_state = "hidden"

    last_row = 1 + len(df)
    refs = {
        col: DataRef(sheet_name, get_column_letter(j), 2, last_row)
        for j, col in enumerate(df.columns, start=1)
    }
    return refs


def write_full_data_sheet(wb: Workbook, full_df: pd.DataFrame) -> None:
    write_data_sheet(wb, "Dữ liệu đã số hóa", full_df, hidden=False)


# ---------------------------------------------------------------------------
# Tầng 1 — "Thống kê tổng hợp"
# ---------------------------------------------------------------------------

@dataclass
class FreqBlockLocation:
    sheet: str
    header_row: int
    first_data_row: int
    last_data_row: int
    label_col: str
    n_col: str
    pct_col: str

    def range(self, col_letter: str) -> str:
        return f"'{self.sheet}'!${col_letter}${self.first_data_row}:${col_letter}${self.last_data_row}"


def _write_freq_block(ws: Worksheet, row: int, column: str, data_ref: DataRef, df: pd.DataFrame) -> tuple[int, FreqBlockLocation]:
    meta = CODEBOOK[column]
    ws.cell(row=row, column=1, value=meta["label"]).font = SECTION_FONT
    ws.cell(row=row, column=2, value=f"[{column}]").font = NOTE_FONT
    row += 1

    if meta["kind"] == "binary":
        valid_n_formula = f'=COUNTIF({data_ref.range},1)+COUNTIF({data_ref.range},0)'
    elif meta["kind"] == "boolean":
        valid_n_formula = f'=COUNTIF({data_ref.range},TRUE)+COUNTIF({data_ref.range},FALSE)'
    else:
        valid_n_formula = f'=COUNTA({data_ref.range})'
    ws.cell(row=row, column=1, value="n hợp lệ")
    ws.cell(row=row, column=2, value=valid_n_formula)
    ws.cell(row=row, column=3, value=f'=COUNTBLANK({data_ref.range})')
    ws.cell(row=row, column=4, value="n hợp lệ / n thiếu (missing) — chỉ để tham khảo, KHÔNG dùng làm mẫu số %").font = NOTE_FONT
    row += 1

    # 26/07: mẫu số % = TỔNG SỐ PHIẾU cố định (literal, = số dòng dữ liệu thật của cột
    # này = len(df)), KHÔNG dùng "n hợp lệ" ở trên nữa — phiếu bỏ trống câu này vẫn tính
    # vào mẫu số 85, không bị loại ra rồi chia trên phần còn lại. Ghi literal (không phải
    # COUNTA sống) vì số dòng dữ liệu đã cố định ngay lúc build sheet này.
    total_n = len(df)
    total_n_cell_row = row
    ws.cell(row=row, column=1, value="N tổng mẫu (mẫu số dùng cho cột %)")
    ws.cell(row=row, column=2, value=total_n)
    total_n_addr = f"$B${total_n_cell_row}"
    row += 1

    header_row = row
    ws.cell(row=row, column=1, value="Lựa chọn").font = HEADER_FONT
    ws.cell(row=row, column=2, value="n").font = HEADER_FONT
    ws.cell(row=row, column=3, value="%").font = HEADER_FONT
    for c in (1, 2, 3):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    row += 1

    first_data_row = row
    rows_def = _fixed_rows_def(meta, column, df)

    for code, label in rows_def:
        ws.cell(row=row, column=1, value=label)
        count_formula = f'={_count_formula(meta, data_ref.range, code)}'
        n_cell = ws.cell(row=row, column=2, value=count_formula)
        ws.cell(row=row, column=3, value=f"=IF({total_n_addr}=0,0,B{row}/{total_n_addr})")
        ws.cell(row=row, column=3).number_format = "0.0%"
        row += 1
    last_data_row = row - 1

    # 26/07 (phản hồi khách — bảng số trần trụi "không thể hiện được ý nghĩa"): thêm
    # data bar sống ngay trong cột % — nhìn là thấy lựa chọn nào áp đảo mà không cần
    # mở sang sheet "Biểu đồ" riêng. end_type="max" (không phải "num"/1) vì vài câu
    # multi-select/combo có thể cộng vượt 100% (has_overlap) — tự co giãn theo giá trị
    # lớn nhất trong chính khối đó thay vì giả định trần cố định 100%.
    if last_data_row >= first_data_row:
        ws.conditional_formatting.add(
            f"C{first_data_row}:C{last_data_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color="638EC6", showValue=True),
        )

    row += 1  # dòng trống ngăn cách khối
    loc = FreqBlockLocation(ws.title, header_row, first_data_row, last_data_row, "A", "B", "C")
    return row, loc


def _write_grouped_binary_block(
    ws: Worksheet, row: int, title: str, note: str | None,
    items: list[tuple[str, str]], data_refs: dict[str, DataRef], df: pd.DataFrame,
) -> tuple[int, FreqBlockLocation]:
    """1 bảng DUY NHẤT cho toàn bộ các cột nhị phân thuộc CÙNG 1 câu multi-select/device-grid
    (`items` = [(tên cột, nhãn lựa chọn), ...]) — mỗi lựa chọn 1 DÒNG n/%, KHÔNG tách riêng
    "Có chọn (n/%) / Không chọn (n/%)" cho từng lựa chọn như trước (phản hồi khách 26/07 tối:
    "làm gì có nhà phân tích thống kê nào làm kiểu đấy" — multi-select phải trình bày % mỗi
    lựa chọn cạnh nhau trong 1 bảng, không phải N bảng nhị phân rời rạc). % vẫn chia cho tổng
    mẫu cố định (len(df)) — có thể cộng vượt 100% vì 1 phiếu chọn được nhiều phương án, đúng
    bản chất multi-select, không phải lỗi."""
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    row += 1
    if note:
        ws.cell(row=row, column=1, value=note).font = NOTE_FONT
        row += 1

    header_row = row
    for col_idx, h in enumerate(("Lựa chọn", "n", "%"), start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1

    total_n = len(df)
    first_data_row = row
    for col_name, label in items:
        ref = data_refs[col_name]
        meta = CODEBOOK[col_name]
        ws.cell(row=row, column=1, value=label)
        true_literal = "TRUE" if meta["kind"] == "boolean" else "1"
        ws.cell(row=row, column=2, value=f"=COUNTIF({ref.range},{true_literal})")
        pct_cell = ws.cell(row=row, column=3, value=f"=IF({total_n}=0,0,B{row}/{total_n})")
        pct_cell.number_format = "0.0%"
        row += 1
    last_data_row = row - 1
    if last_data_row >= first_data_row:
        ws.conditional_formatting.add(
            f"C{first_data_row}:C{last_data_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color="638EC6", showValue=True),
        )
    row += 1
    loc = FreqBlockLocation(ws.title, header_row, first_data_row, last_data_row, "A", "B", "C")
    return row, loc


def _write_matrix_freq_block(
    ws: Worksheet, row: int, qid: str, row_codes: list[str], data_refs: dict[str, DataRef], df: pd.DataFrame,
) -> tuple[int, FreqBlockLocation, list[tuple[str, int, int]]]:
    """1 bảng ma trận DUY NHẤT cho toàn bộ các dòng thuộc CÙNG 1 câu ma trận (Q14/Q32) — mỗi
    việc/vấn đề là 1 DÒNG, mỗi lựa chọn (vợ/chồng/cả hai/khác...) là 1 CẶP cột n/%, thay vì
    18 (hay 8) khối rời rạc mỗi khối 1 dòng như trước."""
    first_meta = CODEBOOK[f"{qid}_{row_codes[0]}"]
    base_label = first_meta["label"].split(" — ", 1)[0]
    options = first_meta["options"] or []

    ws.cell(row=row, column=1, value=f"{base_label} (ma trận — {len(row_codes)} việc/vấn đề)").font = SECTION_FONT
    row += 1
    ws.cell(
        row=row, column=1,
        value="% mỗi ô tính trên 85 phiếu, riêng theo từng dòng (mỗi dòng là 1 việc/vấn đề độc lập).",
    ).font = NOTE_FONT
    row += 1

    header_row = row
    ws.cell(row=row, column=1, value="Việc/vấn đề").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    col_layout: list[tuple[str, int, int]] = []
    next_col = 2
    for opt_code, opt_label in options:
        for suffix, width in ((" (n)", 11), (" (%)", 9)):
            c = ws.cell(row=row, column=next_col, value=f"{opt_label}{suffix}")
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[get_column_letter(next_col)].width = width
            next_col += 1
        col_layout.append((opt_code, next_col - 2, next_col - 1))
    ws.row_dimensions[row].height = 28
    row += 1

    total_n = len(df)
    first_data_row = row
    for row_code in row_codes:
        col = f"{qid}_{row_code}"
        ref = data_refs[col]
        meta = CODEBOOK[col]
        row_label = meta["label"].split(" — ", 1)[-1]
        ws.cell(row=row, column=1, value=row_label)
        for opt_code, n_col, pct_col in col_layout:
            n_formula = f"={_count_formula(meta, ref.range, opt_code)}"
            ws.cell(row=row, column=n_col, value=n_formula)
            col_letter = get_column_letter(n_col)
            pct_cell = ws.cell(row=row, column=pct_col, value=f"=IF({total_n}=0,0,{col_letter}{row}/{total_n})")
            pct_cell.number_format = "0.0%"
        row += 1
    last_data_row = row - 1
    if last_data_row >= first_data_row:
        for _opt_code, _n_col, pct_col in col_layout:
            col_letter = get_column_letter(pct_col)
            ws.conditional_formatting.add(
                f"{col_letter}{first_data_row}:{col_letter}{last_data_row}",
                DataBarRule(start_type="num", start_value=0, end_type="max", color="9DB8D2", showValue=True),
            )
    row += 1
    first_pair = col_layout[0]
    loc = FreqBlockLocation(ws.title, header_row, first_data_row, last_data_row, "A", get_column_letter(first_pair[1]), get_column_letter(first_pair[2]))
    return row, loc, col_layout


def _write_device_grid_freq_block(
    ws: Worksheet, row: int, qid: str, spec: dict[str, Any], data_refs: dict[str, DataRef], df: pd.DataFrame,
) -> tuple[int, FreqBlockLocation]:
    """Device-grid (Q17) — cùng 1 bảng cho mọi thiết bị × người sở hữu, mỗi lựa chọn (vd
    'Điện thoại thông minh — Vợ') là 1 dòng, cộng thêm 1 dòng cho extra_option ('Không ai có
    thiết bị nào'). Dùng lại _write_grouped_binary_block — bản chất mỗi ô sở hữu là 1 cờ nhị
    phân độc lập, giống hệt shape multi-select."""
    items: list[tuple[str, str]] = []
    for row_code in spec["rows"]:
        for col_code in spec["columns"]:
            col_name = f"{qid}_{row_code}_{col_code}"
            label = CODEBOOK[col_name]["label"].split(" — ", 1)[-1]
            items.append((col_name, label))
    extra = spec.get("extra_option")
    if extra:
        col_name = f"{qid}_{extra}"
        label = CODEBOOK[col_name]["label"]
        items.append((col_name, label))
    return _write_grouped_binary_block(ws, row, f"{CODEBOOK[items[0][0]]['label'].split(' — ', 1)[0]} (đa lựa chọn)", None, items, data_refs, df)


def _write_descriptive_block(ws: Worksheet, row: int, column: str, data_ref: DataRef) -> int:
    meta = CODEBOOK[column]
    ws.cell(row=row, column=1, value=meta["label"]).font = SECTION_FONT
    ws.cell(row=row, column=2, value=f"[{column}]").font = NOTE_FONT
    row += 1
    stats = [
        ("n", f"=COUNT({data_ref.range})"),
        ("Trung bình (mean)", f"=AVERAGE({data_ref.range})"),
        ("Trung vị (median)", f"=MEDIAN({data_ref.range})"),
        # Dùng hàm cổ điển (MODE/STDEV/QUARTILE) thay vì bản mới MODE.SNGL/STDEV.S/
        # QUARTILE.INC: openpyxl ghi formula string thô, không tự thêm tiền tố "_xlfn."
        # mà Excel thật đòi hỏi cho hàm ra mắt từ 2010 trở đi -> thiếu tiền tố sẽ ra
        # #NAME? (phát hiện khi khách mở bằng Excel thật, thư viện kiểm chứng lúc build
        # lại hiểu bare name nên không bắt được lỗi này). Hàm cổ điển cho kết quả giống
        # hệt (STDEV=STDEV.S, MODE=MODE.SNGL, QUARTILE=QUARTILE.INC) và chạy được trên
        # mọi bản Excel/WPS Office, an toàn hơn cho khách non-tech.
        ("Yếu vị (mode)", f"=MODE({data_ref.range})"),
        ("Độ lệch chuẩn (SD)", f"=STDEV({data_ref.range})"),
        ("Nhỏ nhất (min)", f"=MIN({data_ref.range})"),
        ("Lớn nhất (max)", f"=MAX({data_ref.range})"),
        ("Tứ phân vị 1 (Q1)", f"=QUARTILE({data_ref.range},1)"),
        ("Tứ phân vị 3 (Q3)", f"=QUARTILE({data_ref.range},3)"),
    ]
    for label, formula in stats:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        row += 1
    return row + 1


def write_frequency_sheet(
    wb: Workbook, data_refs: dict[str, DataRef], df: pd.DataFrame, column_order: list[str]
) -> tuple[dict[str, FreqBlockLocation], dict[str, list[tuple[str, int, int]]]]:
    """Trả về (locations, matrix_layouts) — matrix_layouts chỉ có entry cho Q14/Q32 (ma
    trận nhiều cột lựa chọn/dòng), dùng để vẽ chart nhiều-chuỗi (1 chuỗi/lựa chọn) thay vì
    chart 1-chuỗi thường (xem write_all_charts_sheet). 26/07 đêm — đổi từ chỉ trả `locations`
    sang tuple, vì cần thêm thông tin layout đầy đủ để "vẽ tất cả biểu đồ" (phản hồi khách)
    mà không phá vỡ FreqBlockLocation (giữ nguyên shape đơn giản cho các bảng thường).
    CẢNH BÁO: build_client_report.py (script CŨ, không dùng trong pipeline hiện tại — xem
    build_pillar_report.py) vẫn gọi hàm này theo API CŨ (`locations = write_frequency_sheet
    (...)`), sẽ lỗi nếu chạy lại — chấp nhận được vì script đó đã bị thay thế, không nằm
    trong luồng build chính thức (ghi rõ ở đầu build_pillar_report.py)."""
    ws = wb.create_sheet("Thống kê tổng hợp")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    ws.cell(row=1, column=1, value="TẦNG 1 — THỐNG KÊ TRẢI PHẲNG (toàn bộ biến, % tính trên tổng mẫu, không loại missing khỏi mẫu số)").font = TITLE_FONT
    row = 3

    ws.cell(row=row, column=1, value="Bảng mô tả — 4 biến số liên tục").font = TITLE_FONT
    row += 1
    for col in CONTINUOUS_COLUMNS:
        row = _write_descriptive_block(ws, row, col, data_refs[col])

    ws.cell(row=row, column=1, value="Bảng tần suất — toàn bộ biến categorical/binary/boolean").font = TITLE_FONT
    row += 1

    locations: dict[str, FreqBlockLocation] = {}
    matrix_layouts: dict[str, list[tuple[str, int, int]]] = {}
    written_qids: set[str] = set()
    for col in column_order:
        if col in CONTINUOUS_COLUMNS:
            continue
        qid = GROUPED_QUESTION_COLUMNS.get(col)
        if qid is not None:
            # 26/07 (tối) — cột này thuộc 1 câu multi-select/ma trận/device-grid đã gộp
            # thành 1 khối duy nhất; chỉ ghi khối đó 1 LẦN ở vị trí cột ĐẦU TIÊN của câu
            # này trong column_order, các cột còn lại của cùng câu bị bỏ qua (đã có trong
            # khối vừa ghi).
            if qid in written_qids:
                continue
            written_qids.add(qid)
            if qid in MATRIX_ROW_FIELDS:
                row, loc, col_layout = _write_matrix_freq_block(ws, row, qid, MATRIX_ROW_FIELDS[qid], data_refs, df)
                matrix_layouts[qid] = col_layout
            elif qid in DEVICE_GRID_FIELDS:
                row, loc = _write_device_grid_freq_block(ws, row, qid, DEVICE_GRID_FIELDS[qid], data_refs, df)
            else:
                items = [
                    (f"{qid}_{code}", CODEBOOK[f"{qid}_{code}"]["label"].split(" — ", 1)[-1])
                    for code in MULTI_SELECT_FIELDS[qid]
                ]
                base_label = CODEBOOK[items[0][0]]["label"].split(" — ", 1)[0]
                note = "% mỗi lựa chọn tính trên 85 phiếu — có thể cộng vượt 100% vì 1 phiếu chọn được nhiều phương án."
                row, loc = _write_grouped_binary_block(ws, row, f"{base_label} (đa lựa chọn)", note, items, data_refs, df)
            locations[qid] = loc
            continue
        row, loc = _write_freq_block(ws, row, col, data_refs[col], df)
        locations[col] = loc

    return locations, matrix_layouts


# ---------------------------------------------------------------------------
# Tầng 2 — "Cross-tab" (theo tỉnh, §3.1 mức "theo vùng" chính; theo xã, §3.1 "bảng chi
# tiết phụ")
# ---------------------------------------------------------------------------

# 7 xã theo đúng §3.1 (n khớp thực tế trong output/combined.csv, kiểm tra 25/07): Hàm
# Rồng 25, Mao Sao Phìn 23, Lùng Phình 16, Tả Phìn 10, Bắc Hà 6, Sì Lở Lầu 4, Mã Tra 1.
COMMUNE_LABELS = [
    ("ham-rong", "Hàm Rồng (25)"),
    ("mao-sao-phin", "Mao Sao Phìn (23)"),
    ("lung-phinh", "Lùng Phình (16)"),
    ("ta-phin", "Tả Phìn (10)"),
    ("bac-ha", "Bắc Hà (6)"),
    ("si-lo-lau", "Sì Lở Lầu (4)"),
    ("ma-tra", "Mã Tra (1)"),
]
# §3.1: "Không tính %/so sánh cho Mã Tra (n=1)" — xã duy nhất bị loại khỏi cột %.
NO_PCT_COMMUNES = {"ma-tra"}


@dataclass
class _CrosstabGroup:
    label: str
    criteria: str | None  # None = Toàn thể (không lọc); else f'{ref.range},"{code}"' để nối vào COUNTIFS
    total_formula: str  # mẫu số % = CỠ NHÓM CỐ ĐỊNH (số phiếu thuộc nhóm này), không phải tổng đã trả lời câu hỏi
    show_pct: bool = True


def _crosstab_groups(province_ref: DataRef, commune_ref: DataRef) -> list[_CrosstabGroup]:
    # 26/07: mẫu số % mỗi nhóm = CỠ NHÓM (số phiếu thuộc tỉnh/xã đó, cố định), không phải
    # SUM các hàng đã trả lời câu hỏi trong nhóm (cách cũ bị "phồng" % khi câu đó có phiếu
    # bỏ trống trong nhóm). province/commune luôn được điền cho mọi phiếu (lấy từ manifest,
    # xem scripts/lib/flatten.py) nên COUNTA/COUNTIF trên 2 cột này luôn phản ánh đúng cỡ
    # nhóm thật, không bị lệch vì missing.
    groups = [_CrosstabGroup("Toàn thể", None, f"COUNTA({province_ref.range})")]
    for code, label in (("lao-cai", "Lào Cai"), ("lai-chau", "Lai Châu")):
        groups.append(_CrosstabGroup(label, f'{province_ref.range},"{code}"', f'COUNTIF({province_ref.range},"{code}")'))
    for code, label in COMMUNE_LABELS:
        groups.append(_CrosstabGroup(
            label, f'{commune_ref.range},"{code}"', f'COUNTIF({commune_ref.range},"{code}")',
            show_pct=code not in NO_PCT_COMMUNES,
        ))
    return groups


def write_crosstab_sheet(
    wb: Workbook, data_refs: dict[str, DataRef], df: pd.DataFrame, column_order: list[str]
) -> None:
    ws = wb.create_sheet("Cross-tab")
    ws.column_dimensions["A"].width = 42

    ws.cell(row=1, column=1, value="TẦNG 2 — CROSS-TAB THEO TỈNH + CHI TIẾT THEO XÃ").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="Lào Cai n=58, Lai Châu n=27 — mức 'theo vùng' chính, đủ lớn để so sánh có ý "
              "nghĩa (§3.1). Các cột theo xã bên phải chỉ là bảng chi tiết phụ tham khảo — "
              "xã Mã Tra chỉ có 1 phiếu nên KHÔNG tính %, chỉ hiện số phiếu (n).",
    ).font = NOTE_FONT
    row = 4

    groups = _crosstab_groups(data_refs["province"], data_refs["commune"])
    # Cấp phát cột: Toàn thể/tỉnh có 2 cột (n, %); Mã Tra chỉ 1 cột (n).
    col_layout: list[tuple[_CrosstabGroup, int, int | None]] = []
    next_col = 2
    for g in groups:
        n_col = next_col
        pct_col = next_col + 1 if g.show_pct else None
        col_layout.append((g, n_col, pct_col))
        next_col += 2 if g.show_pct else 1
    for _g, n_col, pct_col in col_layout:
        ws.column_dimensions[get_column_letter(n_col)].width = 11
        if pct_col:
            ws.column_dimensions[get_column_letter(pct_col)].width = 9

    for col in column_order:
        if col in CONTINUOUS_COLUMNS or col in ("province", "commune"):
            continue
        meta = CODEBOOK[col]
        data_ref = data_refs[col]
        ws.cell(row=row, column=1, value=meta["label"]).font = SECTION_FONT
        ws.cell(row=row, column=2, value=f"[{col}]").font = NOTE_FONT
        row += 1

        header_row = row
        ws.cell(row=row, column=1, value="Lựa chọn").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        for g, n_col, pct_col in col_layout:
            c = ws.cell(row=row, column=n_col, value=f"{g.label} (n)")
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            if pct_col:
                c = ws.cell(row=row, column=pct_col, value=f"{g.label} (%)")
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
        row += 1

        rows_def = _fixed_rows_def(meta, col, df)
        first_data_row = row
        for code, label in rows_def:
            ws.cell(row=row, column=1, value=label)
            for g, n_col, _pct_col in col_layout:
                ws.cell(row=row, column=n_col, value=f"={_count_formula(meta, data_ref.range, code, g.criteria)}")
            row += 1
        last_data_row = row - 1

        # 26/07: mẫu số % = g.total_formula (cỡ nhóm cố định — số phiếu thuộc tỉnh/xã đó),
        # KHÔNG phải SUM các hàng đã trả lời câu hỏi trong nhóm (cách cũ, bị "phồng" % khi
        # câu đó có phiếu bỏ trống trong nhóm). Xem _crosstab_groups().
        for r in range(first_data_row, last_data_row + 1):
            for g, n_col, pct_col in col_layout:
                if pct_col is None:
                    continue
                col_letter_n = get_column_letter(n_col)
                cell = ws.cell(row=r, column=pct_col, value=f"=IF({g.total_formula}=0,0,{col_letter_n}{r}/{g.total_formula})")
                cell.number_format = "0.0%"

        row += 1  # dòng trống ngăn cách


# ---------------------------------------------------------------------------
# Charts (Phần A, §11) — tham chiếu trực tiếp vào khối tần suất vừa ghi
# ---------------------------------------------------------------------------

def write_chart_sheet(wb: Workbook, locations: dict[str, FreqBlockLocation], chart_specs: list[dict]) -> None:
    ws = wb.create_sheet("Biểu đồ")
    anchor_row = 1
    for spec in chart_specs:
        col = spec["column"]
        if col not in locations:
            continue
        loc = locations[col]
        chart = BarChart() if spec["kind"] == "bar" else PieChart()
        chart.title = CODEBOOK[col]["label"]
        chart.height = 7
        chart.width = 14
        cats = Reference(wb["Thống kê tổng hợp"], min_col=1, min_row=loc.first_data_row, max_row=loc.last_data_row)
        data = Reference(wb["Thống kê tổng hợp"], min_col=2, min_row=loc.header_row, max_row=loc.last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        if spec["kind"] == "bar":
            chart.type = "bar"  # ngang, dễ đọc tên dài tiếng Việt
        # 26/07 (phản hồi khách): chỉ có cột/lát cắt, không thấy số % rõ ràng — bật nhãn dữ
        # liệu ngay trên biểu đồ (Excel tự hiện đúng % vì cột nguồn đã format "0.0%").
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        ws.add_chart(chart, f"A{anchor_row}")
        anchor_row += 16


# ---------------------------------------------------------------------------
# 26/07 (đêm, phản hồi khách — "bảng có gì thì tưng đó biểu đồ chứ", "text đang bị chồng
# lên chart"): thay cho danh sách chart CHỌN LỌC (write_chart_sheet/PILLAR_CHART_SPECS cũ,
# chỉ vài cột) bằng hàm vẽ chart cho MỌI location được đưa vào — không sót bảng nào. Chiều
# cao/khoảng cách mỗi chart tự co giãn theo SỐ DÒNG dữ liệu thật của đúng bảng đó (bảng 18
# dòng như Q14 cần nhiều chỗ hơn bảng 2 dòng như Q6) — bản cũ dùng bước nhảy cố định 16 dòng
# cho MỌI chart bất kể to nhỏ, khiến chart nhiều lựa chọn (nhãn dài, nhiều thanh) bị bó hẹp/
# đè chữ lên nhau hoặc lấn sang vị trí chart kế tiếp.
# ---------------------------------------------------------------------------

def _chart_height_cm(n_rows: int) -> float:
    """~0.5cm/dòng dữ liệu + phần cố định cho tiêu đề/trục — tối thiểu 5cm (bảng 2 dòng vẫn
    đọc được), tối đa 22cm (bảng 18 dòng như Q14 không phá khổ giấy/màn hình quá mức)."""
    return max(5.0, min(22.0, 2.2 + 0.5 * n_rows))


def _chart_anchor_step(height_cm: float) -> int:
    """Số DÒNG lưới cần nhảy qua để chart kế tiếp không đè lên chart này — quy đổi từ cm
    sang dòng lưới mặc định Excel (~0.53cm/dòng ở zoom 100%), cộng thêm đệm 3 dòng."""
    return max(16, int(height_cm / 0.53) + 3)


def _resolve_chart_title(key: str, override: str | None = None) -> str:
    if override:
        return override
    if key in CODEBOOK:
        return CODEBOOK[key]["label"]
    return GROUP_BASE_LABELS.get(key, key)


def write_all_charts_sheet(
    wb: Workbook,
    sheet_name: str,
    locations: dict[str, FreqBlockLocation],
    matrix_layouts: dict[str, list[tuple[str, int, int]]] | None = None,
    titles: dict[str, str] | None = None,
    exclude: set[str] | None = None,
) -> Worksheet:
    """1 chart cho MỖI key trong `locations` (trừ `exclude`) — mặc định chart 1 chuỗi (cột %
    của bảng); nếu key có mặt trong `matrix_layouts` (Q14/Q32 — nhiều cặp cột n/% cho nhiều
    lựa chọn/dòng), vẽ chart NHIỀU CHUỖI (1 chuỗi/lựa chọn, vd 'Vợ'/'Chồng'/'Cả hai'/'Khác'),
    không chỉ chuỗi đầu tiên."""
    matrix_layouts = matrix_layouts or {}
    titles = titles or {}
    exclude = exclude or set()
    ws = wb.create_sheet(sheet_name)
    anchor_row = 1
    for key, loc in locations.items():
        if key in exclude:
            continue
        n_rows = max(1, loc.last_data_row - loc.first_data_row + 1)
        col_layout = matrix_layouts.get(key)
        n_series = len(col_layout) if col_layout else 1
        if col_layout:
            # 26/07 đêm — bug #4 phát hiện qua ẢNH RENDER (Q14/Q32 — bảng ma trận nhiều lựa
            # chọn/dòng, vd 5 chuỗi Vợ/Chồng/Cả hai/Con cái/Người khác × ~16-18 dòng việc):
            # công thức chiều cao cũ chỉ tính theo SỐ DÒNG, không tính SỐ CHUỖI/dòng — mỗi
            # dòng thực chất có n_series thanh xếp cụm, cần nhiều chỗ hơn hẳn 1 dòng đơn
            # chuỗi. Thiếu chỗ -> nhãn % của nhiều thanh cùng dòng chồng đè lên nhau, đọc
            # không nổi (đúng như khách mô tả "text đang bị chồng lên chart").
            height_cm = max(6.0, min(40.0, 2.2 + n_rows * 0.32 * n_series))
        else:
            height_cm = _chart_height_cm(n_rows)
        chart = BarChart()
        chart.type = "bar"
        chart.title = _resolve_chart_title(key, titles.get(key))
        chart.height = height_cm
        chart.width = 17

        # 26/07 đêm — QUAN TRỌNG: openpyxl's set_categories() chỉ gán category cho các
        # series ĐÃ TỒN TẠI trong chart.series tại thời điểm gọi (lặp self.series bên
        # trong). Gọi set_categories() TRƯỚC add_data() (như bản đầu viết) khiến nó
        # không có series nào để gán -> chart rơi về category mặc định 1,2,3... thay vì
        # nhãn thật ở cột A (đây chính là bug "label bị mất, chỉ còn 7;%;14.1%" phát
        # hiện lúc kiểm tra bằng mắt qua ảnh render, không phải qua recalc). Phải add_data
        # TRƯỚC rồi set_categories SAU.
        if col_layout:
            for _opt_code, _n_col, pct_col in col_layout:
                series_ref = Reference(wb[loc.sheet], min_col=pct_col, min_row=loc.header_row, max_row=loc.last_data_row)
                chart.add_data(series_ref, titles_from_data=True)
            chart.grouping = "clustered"
        else:
            pct_col_idx = 3  # cột "C" — luôn là cột % theo mọi _write_*_block hiện có
            data = Reference(wb[loc.sheet], min_col=pct_col_idx, min_row=loc.header_row, max_row=loc.last_data_row)
            chart.add_data(data, titles_from_data=True)

        cats = Reference(wb[loc.sheet], min_col=1, min_row=loc.first_data_row, max_row=loc.last_data_row)
        chart.set_categories(cats)

        # 26/07 đêm — bug #2 phát hiện qua ảnh render (KHÔNG thấy qua recalc): DataLabelList()
        # mặc định của openpyxl/LibreOffice hiện CẢ tên hạng mục (showCatName) lẫn tên chuỗi
        # (showSerName) CÙNG với giá trị -> nhãn dài như "Không có nhu cầu; %; 55.3%" chồng
        # chữ lên trục và lên nhau khi cột ngắn. Tên hạng mục đã có sẵn ở trục Y, tên chuỗi
        # đã có ở chú giải (legend) — chỉ cần showVal=True, tắt hẳn 3 cờ còn lại.
        # 26/07 đêm (tiếp bug #4): với chart NHIỀU CHUỖI (col_layout — Q14/Q32), dù đã tăng
        # chiều cao, mỗi dòng vẫn có tới 5 nhãn % xếp sát nhau — vẫn chồng chữ nếu bật nhãn
        # từng thanh. Ảnh render cho thấy đây là chỗ RÕ NHẤT khách gặp phải — tắt hẳn nhãn số
        # trên từng thanh cho loại chart này, đọc giá trị qua chú giải màu + trục % thay vì
        # nhãn chữ chi chít; chart 1 chuỗi (đa số các bảng khác) vẫn giữ nhãn vì không chồng.
        if col_layout:
            chart.dataLabels = None
        else:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False
            chart.dataLabels.showSerName = False
            chart.dataLabels.showLegendKey = False
            chart.dataLabels.showPercent = False
        # 26/07 đêm: labels dài tiếng Việt + nhiều thanh -> chữ trục Y dễ bị cắt/đè nếu chart
        # quá thấp so với số dòng; đã bù bằng height_cm ở trên, thêm overlap=-10 cho chart
        # nhiều chuỗi để các cột cạnh nhau tách rõ, không dính vào nhau.
        if col_layout:
            chart.overlap = -10
            chart.gapWidth = 60

        # 26/07 đêm — bug #5/#6 phát hiện SAU KHI khách xác nhận đang xem bằng Excel THẬT
        # (không phải preview nhẹ): (a) mặc định openpyxl để CẢ HAI trục axPos="l" cho chart
        # ngang (type="bar") — LibreOffice tự suy luận đúng hướng nên không lộ lỗi khi render
        # qua nó, nhưng Excel thật hiểu đúng theo XML, 2 trục cùng "trái" đè lên nhau làm mất
        # nhãn category; (b) tickLblPos (có hiện nhãn trục hay không) KHÔNG được đặt — thiếu
        # phần tử này LibreOffice mặc định "nextTo" (hiện), Excel thật lại ẩn đi. Phải đặt
        # TƯỜNG MINH cả 2, không dựa vào default ngầm định (2 phần mềm hiểu default khác nhau).
        chart.x_axis.axPos = "l"
        chart.y_axis.axPos = "b"
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.delete = False
        chart.y_axis.delete = False

        ws.add_chart(chart, f"A{anchor_row}")
        anchor_row += _chart_anchor_step(height_cm)

    # 26/07 đêm — bug #3 phát hiện qua ẢNH RENDER (không thấy qua recalc, không thấy khi mở
    # thường trong Excel — chỉ lộ ra khi In/Xem trước khi in/Xuất PDF): sheet này KHÔNG có ô
    # dữ liệu nào, chỉ có chart THẢ NỔI (floating) — Excel/LibreOffice suy ra "vùng in" mặc
    # định từ vùng CÓ Ô DỮ LIỆU, gần như rỗng ở đây, nên chia trang in theo khổ mặc định rất
    # hẹp, cắt vụn 1 chart lớn thành hàng chục trang in phần lớn TRẮNG TOÀN BỘ. Đặt vùng in +
    # khổ ngang + "vừa 1 trang theo chiều ngang" tường minh để ai in/xuất PDF cũng thấy chart
    # trọn vẹn từng cái, không bị cắt vụn.
    if anchor_row > 1:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_area = f"A1:J{anchor_row}"
    return ws


# ---------------------------------------------------------------------------
# "Đa chiều" (§11) — Q32 (ai quyết định) so theo tỉnh, radar chart NATIVE Excel
# ---------------------------------------------------------------------------

def write_q32_radar(wb: Workbook, df: pd.DataFrame) -> None:
    from .reliability import Q32_POSITIVE_CODES, Q32_ROWS, binary_item_matrix

    item_df = binary_item_matrix(df, "Q32", Q32_ROWS, Q32_POSITIVE_CODES)
    ws = wb["Biểu đồ"]
    anchor_col = 10  # J — tách khỏi các chart Phần A ở cột A

    ws.cell(row=1, column=anchor_col, value="ĐA CHIỀU — Q32 SO THEO TỈNH (% 'vợ' hoặc 'cùng quyết định')").font = TITLE_FONT
    header_row = 2
    ws.cell(row=header_row, column=anchor_col, value="Vấn đề").font = HEADER_FONT
    ws.cell(row=header_row, column=anchor_col).fill = HEADER_FILL
    for j, label in enumerate(("Lào Cai (%)", "Lai Châu (%)")):
        c = ws.cell(row=header_row, column=anchor_col + 1 + j, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    row = header_row + 1
    first_data_row = row
    for row_code in Q32_ROWS:
        col = f"Q32_{row_code}"
        label = CODEBOOK[col]["label"].split(" — ", 1)[-1]
        ws.cell(row=row, column=anchor_col, value=label)
        for j, prov_code in enumerate(("lao-cai", "lai-chau")):
            mask = (df["province"] == prov_code).to_numpy()
            pct = round(float(item_df.loc[mask, row_code].mean() * 100), 1) if mask.any() else 0.0
            ws.cell(row=row, column=anchor_col + 1 + j, value=pct)
        row += 1
    last_data_row = row - 1

    for col_letter, width in [(get_column_letter(anchor_col), 34), (get_column_letter(anchor_col + 1), 14), (get_column_letter(anchor_col + 2), 14)]:
        ws.column_dimensions[col_letter].width = width

    chart = RadarChart()
    chart.type = "filled"
    chart.title = "Ai quyết định 8 vấn đề trong gia đình — so theo tỉnh (%)"
    chart.height = 12
    chart.width = 18
    cats = Reference(ws, min_col=anchor_col, min_row=first_data_row, max_row=last_data_row)
    data = Reference(ws, min_col=anchor_col + 1, max_col=anchor_col + 2, min_row=header_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{get_column_letter(anchor_col)}{last_data_row + 3}")
