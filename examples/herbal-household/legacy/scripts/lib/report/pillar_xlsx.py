"""Sheet Excel cho 4 trụ cột + chỉ số vai trò (docs/implement-plan-statistics-and-client-report.md,
bản 26/07) — tái dùng đúng phong cách "sống bằng công thức Excel gốc" đã có ở xlsx_writer.py
(COUNTIF/COUNTIFS tham chiếu sheet "Dữ liệu (ẩn danh)"), không dán số tĩnh trừ 2 chỗ đã nêu
rõ lý do: (1) Q30×Q8 cross-tab dùng cột phụ trợ đã nổ sẵn (Q30_is_<khâu>, xem
`add_derived_columns`) nên vẫn SỐNG qua các cột phụ đó; (2) văn bản SWOT (Pillar D) là tổng
hợp câu chữ, không phải bản thân 1 phép đếm — số trong câu là ảnh chụp tại thời điểm build,
ghi rõ trong sheet.
"""

from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .codebook import build_codebook
from .frequency import category_rows
from .pillars import (
    HIGH_VALUE_NODES,
    Q14_VALUE_CHAIN_POSITIVE_CODES,
    Q14_VALUE_CHAIN_ROWS,
    Q30_NODES,
)
from .reliability import Q32_POSITIVE_CODES, Q32_ROWS
from .reliability_sheet import _composite_formula
from .xlsx_writer import (
    CODEBOOK,
    DataRef,
    FreqBlockLocation,
    HEADER_FILL,
    HEADER_FONT,
    NOTE_FONT,
    SECTION_FONT,
    TITLE_FONT,
    _count_formula,
    _fixed_rows_def,
    _write_descriptive_block,
    _write_freq_block,
    _write_grouped_binary_block,
)

_ = build_codebook  # đảm bảo import không bị linter dọn nhầm (dùng CODEBOOK từ xlsx_writer)


def add_derived_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Cột phụ trợ CHỈ để dựng công thức Excel sống cho Pillar B/C — không đổi combined.csv
    gốc, chỉ thêm vào bản dùng để ghi sheet 'Dữ liệu (ẩn danh)'. Mỗi cột là 1 phép suy ra từ
    Q30 (combo-aware), cùng tinh thần với age_bracket/education_grade_bracket đã có sẵn
    (cột dẫn xuất ngồi cạnh dữ liệu thô để công thức tham chiếu được)."""
    out = df.copy()

    def _has_node(value: Any, node_code: str) -> bool | None:
        if pd.isna(value):
            return None
        parts = value.split("+") if isinstance(value, str) and "+" in value else [value]
        return node_code in parts

    for node_code, _label in Q30_NODES:
        out[f"Q30_is_{node_code}"] = out["Q30"].apply(lambda v, nc=node_code: _has_node(v, nc))

    def _high_value(value: Any) -> bool | None:
        if pd.isna(value):
            return None
        parts = value.split("+") if isinstance(value, str) and "+" in value else [value]
        return any(p in HIGH_VALUE_NODES for p in parts)

    out["Q30_nhom_khau_cao_gia"] = out["Q30"].apply(_high_value)

    # 26/07 (tối) — cột phụ trợ boolean cho cross-tab "lãnh đạo × Q30" (Pillar B, docs
    # implement plan §8): TRUE nếu chọn ít nhất 1 trong 3 vai trò lãnh đạo ở Q33 (ban chủ
    # nhiệm HTX/nhóm SX/quản lý rừng), cùng cách tính OR đã dùng cho Q30_nhom_khau_cao_gia
    # ở trên — cần cột TRUE/FALSE thật (không phải 0/1) để _write_crosstab_block dùng đúng
    # nhánh literal Excel TRUE/FALSE (xem ghi chú trong hàm đó).
    leadership_cols = ["Q33_ban_chu_nhiem_htx", "Q33_nhom_san_xuat", "Q33_quan_ly_rung"]
    out["Q33_co_vai_tro_lanh_dao"] = (out[leadership_cols] == 1).any(axis=1)
    return out


DERIVED_COLUMN_LABELS = {
    **{f"Q30_is_{code}": f"[phụ trợ] Q30 có khâu — {label}" for code, label in Q30_NODES},
    "Q30_nhom_khau_cao_gia": "[phụ trợ] Có tham gia khâu thương mại/tiêu thụ (nhóm 'cao giá')",
    "Q33_co_vai_tro_lanh_dao": "[phụ trợ] Có ít nhất 1 vai trò lãnh đạo (Q33)",
}


def _fmt_pct(n_formula_inner: str, total_formula: str, row: int, col_letter: str) -> str:
    return f"=IF({total_formula}=0,0,{col_letter}{row}/{total_formula})"


def _write_crosstab_block(
    ws, row: int, title: str, note: str | None,
    value_col: str, group_col: str, group_defs: list[tuple[str, Any]],
    df: pd.DataFrame, data_refs: dict[str, DataRef],
) -> int:
    """Bảng so 1 biến (value_col, hàng) theo nhiều nhóm (group_col == code, cột n/%), mẫu số
    = CỠ NHÓM (không phải 85 cố định) — đúng quy ước cross-tab đã có ở crosstab.py."""
    meta = CODEBOOK[value_col]
    value_ref = data_refs[value_col]
    group_ref = data_refs[group_col]

    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    row += 1
    if note:
        ws.cell(row=row, column=1, value=note).font = NOTE_FONT
        row += 1

    ws.cell(row=row, column=1, value="Lựa chọn").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    col_layout = []
    next_col = 2
    for label, _code in group_defs:
        for suffix, width in ((" (n)", 13), (" (%)", 13)):
            c = ws.cell(row=row, column=next_col, value=f"{label}{suffix}")
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            # 26/07: nhãn nhóm dài (vd "Dưới 1 năm kinh nghiệm (n)") tràn sang ô bên
            # cạnh nếu không bọc chữ — bật wrap_text + tăng chiều cao dòng tiêu đề thay
            # vì kéo cột quá rộng (cột rộng làm data bar mất tác dụng trực quan).
            c.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[get_column_letter(next_col)].width = width
            next_col += 1
        col_layout.append((label, next_col - 2, next_col - 1))
    ws.row_dimensions[row].height = 30
    row += 1

    rows_def = _fixed_rows_def(meta, value_col, df)
    first_data_row = row
    for v_code, v_label in rows_def:
        ws.cell(row=row, column=1, value=v_label)
        for (label, g_code), (_l, n_col, pct_col) in zip(group_defs, col_layout):
            if isinstance(g_code, bool):
                extra = f'{group_ref.range},{"TRUE" if g_code else "FALSE"}'
                total_formula = f'COUNTIF({group_ref.range},{"TRUE" if g_code else "FALSE"})'
            else:
                # "=" ở đầu literal — tránh Excel/LibreOffice hiểu nhầm mã nhóm bắt đầu
                # bằng <,>,= (vd "<1", ">=1") thành TOÁN TỬ SO SÁNH thay vì so khớp văn
                # bản (xem giải thích đầy đủ ở xlsx_writer._criteria_literal).
                extra = f'{group_ref.range},"={g_code}"'
                total_formula = f'COUNTIF({group_ref.range},"={g_code}")'
            n_formula = f"={_count_formula(meta, value_ref.range, v_code, extra)}"
            ws.cell(row=row, column=n_col, value=n_formula)
            col_letter = get_column_letter(n_col)
            pct_cell = ws.cell(row=row, column=pct_col, value=_fmt_pct("", total_formula, row, col_letter))
            pct_cell.number_format = "0.0%"
        row += 1
    last_data_row = row - 1
    # Data bar riêng cho từng cột % (mỗi nhóm 1 màu nhạt hơn, tự co giãn theo max của
    # đúng cột đó) — cùng lý do đã ghi ở xlsx_writer._write_freq_block.
    if last_data_row >= first_data_row:
        for _label, _n_col, pct_col in col_layout:
            col_letter = get_column_letter(pct_col)
            ws.conditional_formatting.add(
                f"{col_letter}{first_data_row}:{col_letter}{last_data_row}",
                DataBarRule(start_type="num", start_value=0, end_type="max", color="9DB8D2", showValue=True),
            )
    return row + 1


def _write_positive_share_block(
    ws, row: int, title: str, note: str,
    pairs: list[dict], data_refs: dict[str, DataRef], df: pd.DataFrame,
    lam_meta_by_col: dict[str, tuple[set[str], str]], quyet_meta_by_col: dict[str, tuple[set[str], str]],
) -> int:
    """Bảng '% dòng nào tính là 'vợ có tham gia/quyết định'' cho từng cặp Ai làm/Ai quyết —
    tái dùng _count_formula cho từng positive code rồi cộng dồn (giống cách _count_formula
    tự xử lý combo cho 1 code, ở đây cộng thêm qua NHIỀU code dương)."""
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value=note).font = NOTE_FONT
    row += 1
    headers = ["Chủ đề", "Ai LÀM việc này (%)", "Ai QUYẾT ĐỊNH việc này (%)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    total_n = len(df)
    first_data_row = row
    for p in pairs:
        ws.cell(row=row, column=1, value=p["title"])
        for j, (col, (codes, _lbl)) in enumerate(((p["lam_col"], lam_meta_by_col[p["lam_col"]]), (p["quyet_col"], quyet_meta_by_col[p["quyet_col"]])), start=2):
            ref = data_refs[col]
            meta = CODEBOOK[col]
            terms = [_count_formula(meta, ref.range, code) for code in sorted(codes)]
            formula = f"=({'+'.join(terms)})/{total_n}"
            cell = ws.cell(row=row, column=j, value=formula)
            cell.number_format = "0.0%"
        row += 1
    last_data_row = row - 1
    if last_data_row >= first_data_row:
        for col_letter in ("B", "C"):
            ws.conditional_formatting.add(
                f"{col_letter}{first_data_row}:{col_letter}{last_data_row}",
                DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="55A868", showValue=True),
            )
    return row + 1


def _write_ownership_vs_usage_block(
    ws, row: int, title: str, note: str,
    own_col: str, own_label: str, usage_items: list[tuple[str, str]],
    data_refs: dict[str, DataRef], df: pd.DataFrame,
) -> int:
    """Bảng '% sở hữu thiết bị' so với '% dùng thiết bị đó cho mục đích kinh tế' — KHÁC
    _write_positive_share_block (không phải cặp lam/quyết cùng chủ đề, mà 1 cột sở hữu so
    với NHIỀU cột sử dụng) và khác _write_crosstab_block (không chia nhóm sở hữu có/không —
    tỷ lệ sở hữu quá lệch, 81/85, chia nhóm sẽ có 1 nhóm chỉ 4 phiếu — xem
    pillars.device_ownership_vs_usage)."""
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value=note).font = NOTE_FONT
    row += 1
    headers = ["", "% trên 85 phiếu"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    total_n = len(df)
    first_data_row = row

    own_ref = data_refs[own_col]
    own_meta = CODEBOOK[own_col]
    ws.cell(row=row, column=1, value=own_label)
    cell = ws.cell(row=row, column=2, value=f"={_count_formula(own_meta, own_ref.range, 1)}/{total_n}")
    cell.number_format = "0.0%"
    row += 1

    for label, col in usage_items:
        ref = data_refs[col]
        meta = CODEBOOK[col]
        ws.cell(row=row, column=1, value=f"  ↳ {label}")
        cell = ws.cell(row=row, column=2, value=f"={_count_formula(meta, ref.range, 1)}/{total_n}")
        cell.number_format = "0.0%"
        row += 1
    last_data_row = row - 1
    ws.conditional_formatting.add(
        f"B{first_data_row}:B{last_data_row}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="DD8452", showValue=True),
    )
    return row + 1


def write_pillar_a_sheet(wb: Workbook, df: pd.DataFrame, data_refs: dict[str, DataRef]) -> dict[str, FreqBlockLocation]:
    from .pillars import Q7_OPTIONS

    ws = wb.create_sheet("A. Thị trường")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    ws.cell(row=1, column=1, value="A. THỊ TRƯỜNG & MỨC ĐỘ GẮN BÓ VỚI CÂY DƯỢC LIỆU").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Hộ phụ thuộc thị trường dược liệu đến đâu, và gắn bó lâu hơn có đi cùng phụ thuộc nhiều hơn không.").font = NOTE_FONT
    row = 4
    locations: dict[str, FreqBlockLocation] = {}

    # 26/07 (tối, phản hồi khách — "gộp vào 1 biểu đồ", không tách riêng "có chọn/không
    # chọn" cho từng lựa chọn của 1 câu multi-select): 1 bảng duy nhất cho toàn bộ Q7,
    # thay vì 5 khối rời rạc mỗi khối 1 nguồn thu nhập.
    q7_items = [(f"Q7_{opt}", CODEBOOK[f"Q7_{opt}"]["label"].split(" — ", 1)[-1]) for opt in Q7_OPTIONS]
    row, loc = _write_grouped_binary_block(
        ws, row, "Q7 – Nguồn thu nhập chính (đa lựa chọn)",
        "% mỗi nguồn tính trên 85 phiếu — có thể cộng vượt 100% vì 1 hộ có thể có nhiều nguồn thu nhập.",
        q7_items, data_refs, df,
    )
    locations["Q7"] = loc
    row, loc = _write_freq_block(ws, row, "Q8", data_refs["Q8"], df)
    locations["Q8"] = loc
    row = _write_descriptive_block(ws, row, "Q9_derived_years_exp", data_refs["Q9_derived_years_exp"])

    group_defs = [("Dưới 1 năm kinh nghiệm", "<1"), ("Từ 1 năm trở lên", ">=1")]
    row = _write_crosstab_block(
        ws, row, "Kinh nghiệm trồng dược liệu × Tỷ lệ thu nhập từ dược liệu (Q8)",
        "Mẫu số mỗi cột = số phiếu thuộc đúng nhóm kinh nghiệm đó (không phải 85 cố định).",
        "Q8", "experience_years_bracket", group_defs, df, data_refs,
    )
    return locations


def write_pillar_b_sheet(wb: Workbook, df: pd.DataFrame, data_refs: dict[str, DataRef]) -> FreqBlockLocation:
    ws = wb.create_sheet("B. Chuỗi giá trị")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    ws.cell(row=1, column=1, value="B. VỊ TRÍ TRONG CHUỖI GIÁ TRỊ DƯỢC LIỆU").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Phụ nữ tham gia nhiều nhất ở khâu nào, và 'ai làm' có khớp với 'ai quyết' hay không.").font = NOTE_FONT
    row = 4

    row, q30_loc = _write_freq_block(ws, row, "Q30", data_refs["Q30"], df)

    group_defs = [(label, True) for _code, label in [(c, l) for c, l in Q30_NODES]]
    # nhóm theo TỪNG khâu (cột phụ trợ Q30_is_<khâu>) — dùng field riêng cho từng khâu
    for node_code, node_label in Q30_NODES:
        row = _write_crosstab_block(
            ws, row, f"Trong số phiếu có tham gia khâu '{node_label}' — Tỷ lệ thu nhập từ dược liệu (Q8)",
            None, "Q8", f"Q30_is_{node_code}", [("Có tham gia khâu này", True)], df, data_refs,
        )

    from .pillars import pillar_b as _pillar_b_calc
    pairs = _pillar_b_calc(df)["lam_vs_quyet"]
    lam_meta = {
        "Q14_lien_he_tieu_thu": ({"vo", "ca_hai"}, ""),
        "Q14_quan_ly_chi_tieu": ({"vo", "ca_hai"}, ""),
        "Q13": ({"vo", "ca_hai"}, ""),
    }
    quyet_meta = {
        "Q32_chon_ban": (Q32_POSITIVE_CODES, ""),
        "Q32_su_dung_thu_nhap": (Q32_POSITIVE_CODES, ""),
        "Q32_chon_cay_trong": (Q32_POSITIVE_CODES, ""),
    }
    row = _write_positive_share_block(
        ws, row, "Ai LÀM việc này so với Ai QUYẾT ĐỊNH việc này (cùng chủ đề)",
        "% tính trên 85 phiếu, 'vợ' hoặc 'cả hai'/'cùng quyết định' được tính là có tham gia/quyết định.",
        pairs, data_refs, df, lam_meta, quyet_meta,
    )

    # ------------------------------------------------------------------
    # Bổ sung 26/07 (tối) — gắn học thức/lãnh đạo/đi lại/thiết bị vào câu chuyện Pillar B
    # (docs/implement-plan-statistics-and-client-report.md §8). Các biến này vẫn có đủ ở
    # sheet "Thống kê trải phẳng"; đây là góc nhìn THÊM, không thay thế.
    # ------------------------------------------------------------------
    ws.cell(row=row, column=1, value="Học thức, vai trò lãnh đạo, đi lại độc lập và thiết bị số — có đi cùng tham gia khâu 'cao giá' hơn không").font = TITLE_FONT
    row += 1

    from .pillars import EDUCATION_GROUPS, MOBILITY_GROUPS

    # pillars.py dùng quy ước (code, label); _write_crosstab_block cần (label, code) — đảo
    # lại ở đây, KHÔNG đổi quy ước gốc trong pillars.py (đang dùng đúng chỗ khác trong module đó).
    education_group_defs = [(label, code) for code, label in EDUCATION_GROUPS]
    mobility_group_defs = [(label, code) for code, label in MOBILITY_GROUPS]

    row = _write_crosstab_block(
        ws, row, "Học thức (Q5) × Khâu tham gia chuỗi giá trị (Q30)",
        "Học vấn cao hơn có đi cùng tham gia khâu thương mại/tiêu thụ (khâu gần khách hàng hơn) nhiều hơn không.",
        "Q30", "education_grade_bracket", education_group_defs, df, data_refs,
    )
    row = _write_crosstab_block(
        ws, row, "Có vai trò lãnh đạo nhóm SX/HTX/quản lý rừng (Q33) × Khâu tham gia (Q30)",
        "Mẫu nhóm 'có vai trò lãnh đạo' rất nhỏ (n~5/85) — số liệu chỉ mang tính gợi ý, không kết luận.",
        "Q30", "Q33_co_vai_tro_lanh_dao", [("Có vai trò lãnh đạo", True), ("Không có vai trò lãnh đạo", False)], df, data_refs,
    )
    row = _write_crosstab_block(
        ws, row, "Có xe máy riêng tự đi lại (Q25) × Khâu tham gia (Q30)",
        "Nhiều khâu (đặc biệt thương mại/tiêu thụ) cần di chuyển đến chợ/điểm thu mua — đi lại độc lập có đi cùng tham gia khâu đó nhiều hơn không.",
        "Q30", "Q25", mobility_group_defs, df, data_refs,
    )
    row = _write_ownership_vs_usage_block(
        ws, row, "Sở hữu thiết bị số so với dùng thiết bị đó cho mục đích kinh tế",
        "Sở hữu điện thoại thông minh (vợ) không chia thành 2 nhóm so sánh vì quá lệch (81/85 có) — thay vào đó so trực tiếp tỷ lệ sở hữu với tỷ lệ dùng thực tế cho buôn bán/quảng bá, khoảng cách giữa 2 số này mới là điều đáng chú ý.",
        "Q17_dien_thoai_vo", "Sở hữu điện thoại thông minh (vợ)",
        [("Dùng điện thoại để giao dịch/bán hàng", "Q18_giao_dich_ban_hang"),
         ("Dùng điện thoại để quảng bá sản phẩm", "Q19_quang_ba"),
         ("Bán hàng online", "Q19_ban_hang_online")],
        data_refs, df,
    )
    return q30_loc


def write_pillar_c_sheet(wb: Workbook, df: pd.DataFrame, data_refs: dict[str, DataRef]) -> dict[str, FreqBlockLocation]:
    from .pillars import Q22B_OPTIONS, Q28_OPTIONS

    ws = wb.create_sheet("C. Rào cản")
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    ws.cell(row=1, column=1, value="C. RÀO CẢN SẢN XUẤT").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Rào cản nào phổ biến nhất, và có giữ chân phụ nữ ở khâu thấp giá không.").font = NOTE_FONT
    row = 4
    locations: dict[str, FreqBlockLocation] = {}

    # 26/07 (tối, phản hồi khách — "gộp vào 1 biểu đồ"): 1 bảng duy nhất cho Q28 và 1 bảng
    # cho Q22b, thay vì mỗi lựa chọn 1 khối "có chọn/không chọn" riêng.
    q28_items = [(f"Q28_{opt}", CODEBOOK[f"Q28_{opt}"]["label"].split(" — ", 1)[-1]) for opt in Q28_OPTIONS]
    row, loc = _write_grouped_binary_block(
        ws, row, "Q28 – Khó khăn khi trồng/kinh doanh dược liệu (đa lựa chọn)",
        "% mỗi khó khăn tính trên 85 phiếu — có thể cộng vượt 100% vì 1 phiếu có thể gặp nhiều khó khăn.",
        q28_items, data_refs, df,
    )
    locations["Q28"] = loc
    q22b_items = [(f"Q22b_{opt}", CODEBOOK[f"Q22b_{opt}"]["label"].split(" — ", 1)[-1]) for opt in Q22B_OPTIONS]
    row, loc = _write_grouped_binary_block(
        ws, row, "Q22b – Lý do chưa vay vốn (đa lựa chọn)",
        "% mỗi lý do tính trên 85 phiếu.",
        q22b_items, data_refs, df,
    )
    locations["Q22b"] = loc

    ws.cell(row=row, column=1, value="Rào cản theo tỉnh và theo nhóm khâu tham gia").font = TITLE_FONT
    row += 1
    for opt in Q28_OPTIONS:
        col = f"Q28_{opt}"
        label = CODEBOOK[col]["label"].split(" — ", 1)[-1]
        row = _write_crosstab_block(
            ws, row, f"{label} — theo tỉnh", None,
            col, "province", [("Lào Cai", "lao-cai"), ("Lai Châu", "lai-chau")], df, data_refs,
        )
        row = _write_crosstab_block(
            ws, row, f"{label} — theo nhóm khâu tham gia", None,
            col, "Q30_nhom_khau_cao_gia",
            [("Có tham gia khâu thương mại/tiêu thụ", True), ("Chỉ khâu sản xuất/thu hái/chế biến", False)],
            df, data_refs,
        )

    # 26/07 (tối) — theo dân tộc (Q4). KHÁCH YÊU CẦU KHÔNG GỘP tên dân tộc cụ thể lại (xem
    # docs/client-feedback-2026-07-22-extraction-rules.md §2.4, nhắc lại tối 26/07) — dùng
    # đúng tên quan sát được trong dữ liệu (category_rows), không tự bó "Kinh/DTTS" hay bất
    # kỳ gộp nhóm nào khác. Vài nhóm rất nhỏ (Tày n=4, Nùng n=2) — vẫn hiện riêng theo đúng
    # yêu cầu, ghi rõ caveat mẫu nhỏ ngay trong tiêu đề khối.
    ws.cell(row=row, column=1, value="Rào cản theo dân tộc (Q4 — giữ nguyên từng dân tộc, không gộp)").font = TITLE_FONT
    row += 1
    ws.cell(
        row=row, column=1,
        value="Vài nhóm dân tộc trong mẫu này rất nhỏ (Kinh n=4, Tày n=4, Nùng n=2) — % ở các nhóm đó chỉ mang tính tham khảo, không đại diện thống kê.",
    ).font = NOTE_FONT
    row += 1
    ethnicity_defs = [(label, code) for code, label in category_rows("Q4", df["Q4"])]
    for opt in Q28_OPTIONS:
        col = f"Q28_{opt}"
        label = CODEBOOK[col]["label"].split(" — ", 1)[-1]
        row = _write_crosstab_block(
            ws, row, f"{label} — theo dân tộc", None,
            col, "Q4", ethnicity_defs, df, data_refs,
        )
    return locations


def write_pillar_d_sheet(wb: Workbook, df: pd.DataFrame, data_refs: dict[str, DataRef], swot: dict[str, list[str]]) -> dict[str, FreqBlockLocation]:
    from .pillars import Q11_OPTIONS, Q21B_OPTIONS, Q22A_OPTIONS

    ws = wb.create_sheet("D. Chính sách")
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    ws.cell(row=1, column=1, value="D. MÔI TRƯỜNG CHÍNH SÁCH/THỂ CHẾ").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Kênh hỗ trợ nào có tác dụng, kênh nào chưa tiếp cận được và vì sao.").font = NOTE_FONT
    row = 4
    locations: dict[str, FreqBlockLocation] = {}

    # 26/07 (tối, phản hồi khách — "gộp vào 1 biểu đồ", không tách "có chọn/không chọn"
    # từng lựa chọn): 1 bảng duy nhất cho mỗi câu multi-select (Q22a/Q21b/Q11).
    q22a_items = [(f"Q22a_{opt}", CODEBOOK[f"Q22a_{opt}"]["label"].split(" — ", 1)[-1]) for opt in Q22A_OPTIONS]
    row, loc = _write_grouped_binary_block(
        ws, row, "Q22a – Vay vốn sản xuất/kinh doanh (đa lựa chọn)",
        "% mỗi kênh vay vốn tính trên 85 phiếu — có thể cộng vượt 100% vì 1 hộ có thể vay từ nhiều nguồn.",
        q22a_items, data_refs, df,
    )
    locations["Q22a"] = loc
    row, loc = _write_freq_block(ws, row, "Q23", data_refs["Q23"], df)
    locations["Q23"] = loc
    row, loc = _write_freq_block(ws, row, "Q21a", data_refs["Q21a"], df)
    locations["Q21a"] = loc
    q21b_items = [(f"Q21b_{opt}", CODEBOOK[f"Q21b_{opt}"]["label"].split(" — ", 1)[-1]) for opt in Q21B_OPTIONS]
    row, loc = _write_grouped_binary_block(
        ws, row, "Q21b – Nội dung tập huấn (đa lựa chọn)",
        "% mỗi nội dung tính trên 85 phiếu.",
        q21b_items, data_refs, df,
    )
    locations["Q21b"] = loc
    q11_items = [(f"Q11_{opt}", CODEBOOK[f"Q11_{opt}"]["label"].split(" — ", 1)[-1]) for opt in Q11_OPTIONS]
    row, loc = _write_grouped_binary_block(
        ws, row, "Q11 – Hội đoàn thể tham gia (đa lựa chọn)",
        "% mỗi tổ chức tính trên 85 phiếu — có thể cộng vượt 100% vì 1 người có thể tham gia nhiều tổ chức.",
        q11_items, data_refs, df,
    )
    locations["Q11"] = loc

    # 26/07 (tối) — kênh vay vốn theo dân tộc (Q4, KHÔNG gộp — cùng lý do đã ghi ở Pillar C).
    ws.cell(row=row, column=1, value="Tiếp cận kênh vay vốn theo dân tộc (Q4 — giữ nguyên từng dân tộc, không gộp)").font = TITLE_FONT
    row += 1
    ws.cell(
        row=row, column=1,
        value="Vài nhóm dân tộc trong mẫu này rất nhỏ (Kinh n=4, Tày n=4, Nùng n=2) — % ở các nhóm đó chỉ mang tính tham khảo, không đại diện thống kê.",
    ).font = NOTE_FONT
    row += 1
    ethnicity_defs = [(label, code) for code, label in category_rows("Q4", df["Q4"])]
    for opt in Q22A_OPTIONS:
        col = f"Q22a_{opt}"
        label = CODEBOOK[col]["label"].split(" — ", 1)[-1]
        row = _write_crosstab_block(
            ws, row, f"{label} — theo dân tộc", None,
            col, "Q4", ethnicity_defs, df, data_refs,
        )

    ws.cell(row=row, column=1, value="TỔNG HỢP SWOT — MÔI TRƯỜNG CHÍNH SÁCH").font = TITLE_FONT
    row += 1
    ws.cell(
        row=row, column=1,
        value="Văn bản tổng hợp TĨNH (không phải công thức sống) — số liệu trích từ các khối "
              "tần suất bên trên tại thời điểm build. Chạy lại script nếu dữ liệu review đổi.",
    ).font = NOTE_FONT
    row += 2

    swot_labels = [
        ("strengths", "ĐIỂM MẠNH"), ("weaknesses", "ĐIỂM YẾU"),
        ("opportunities", "CƠ HỘI"), ("threats", "THÁCH THỨC"),
    ]
    for key, label in swot_labels:
        ws.cell(row=row, column=1, value=label).font = SECTION_FONT
        row += 1
        for bullet in swot[key]:
            ws.cell(row=row, column=1, value=f"• {bullet}")
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1

    return locations


def write_pillar_e_sheet(wb: Workbook, df: pd.DataFrame, data_refs: dict[str, DataRef]) -> dict[str, Any]:
    from .pillars import pillar_e as _pillar_e_calc

    result = _pillar_e_calc(df)
    r14, r32 = result["Q14"], result["Q32"]

    ws = wb.create_sheet("E. Chỉ số vai trò")
    ws.cell(row=1, column=1, value="E. CHỈ SỐ VAI TRÒ TRONG CHUỖI GIÁ TRỊ DƯỢC LIỆU (CRONBACH'S ALPHA)").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="Kiểm tra 1 nhóm câu hỏi có đủ nhất quán để gộp thành 1 chỉ số duy nhất không. "
              "Mốc tham khảo: alpha ≥ 0.7 (Nunnally). Khác bản gốc: Q14 chỉ lấy 8 việc liên quan "
              "trực tiếp sản xuất/thương mại dược liệu, bỏ việc nhà thuần tuý.",
    ).font = NOTE_FONT
    ws.cell(row=3, column=1, value="Alpha là số tĩnh (quyết định có tạo chỉ số hay không). Cột 'Chỉ số (%)' từng phiếu LÀ công thức Excel sống.").font = NOTE_FONT
    row = 5

    specs = {"Q14": (Q14_VALUE_CHAIN_ROWS, Q14_VALUE_CHAIN_POSITIVE_CODES), "Q32": (Q32_ROWS, Q32_POSITIVE_CODES)}
    for qid, r in (("Q14", r14), ("Q32", r32)):
        ws.cell(row=row, column=1, value=r.label).font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="Số dòng (item)")
        ws.cell(row=row, column=2, value=r.n_items)
        row += 1
        ws.cell(row=row, column=1, value="Cronbach's alpha")
        ws.cell(row=row, column=2, value=round(r.alpha, 3) if r.alpha is not None else "Không tính được")
        row += 1
        ws.cell(row=row, column=1, value="Quyết định")
        ws.cell(row=row, column=2, value="ĐẠT ngưỡng — tạo chỉ số tổng hợp" if r.composite is not None else "CHƯA đạt ngưỡng — giữ nguyên từng dòng")
        row += 2

        if r.composite is not None:
            ws.cell(row=row, column=1, value="record_id").font = HEADER_FONT
            ws.cell(row=row, column=1).fill = HEADER_FILL
            ws.cell(row=row, column=2, value="Chỉ số (%)").font = HEADER_FONT
            ws.cell(row=row, column=2).fill = HEADER_FILL
            ws.cell(row=row, column=3, value="Tỉnh").font = HEADER_FONT
            ws.cell(row=row, column=3).fill = HEADER_FILL
            first_data_row = row + 1
            record_ids = df["record_id"]
            provinces = df["province"]
            rows_def, positive_codes = specs[qid]
            base_ref = data_refs[f"{qid}_{rows_def[0]}"]
            for pos, idx in enumerate(r.composite.index):
                row += 1
                record_row = base_ref.first_row + pos
                ws.cell(row=row, column=1, value=record_ids.loc[idx])
                cell = ws.cell(row=row, column=2, value=_composite_formula(data_refs, qid, rows_def, positive_codes, record_row))
                cell.number_format = "0.0"
                ws.cell(row=row, column=3, value=provinces.loc[idx])
            last_data_row = row
            ws.conditional_formatting.add(
                f"B{first_data_row}:B{last_data_row}",
                DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="8172B2", showValue=True),
            )
            row += 1
            ws.cell(row=row, column=1, value="Trung bình toàn thể")
            avg_cell = ws.cell(row=row, column=2, value=f"=AVERAGE(B{first_data_row}:B{last_data_row})")
            avg_cell.number_format = "0.0"
            row += 2

            # 26/07 (tối, docs implement plan §8) — chỉ số theo nhóm tuổi kết hôn (Q6):
            # kết hôn sớm có đi cùng vai trò/tiếng nói quyết định thấp hơn không. Sống bằng
            # AVERAGEIFS — dùng ĐÚNG cột B{first_data_row}:{last_data_row} vừa ghi ở trên
            # (cùng thứ tự phiếu với mg_ref bên dưới vì cả 2 đều lấy từ cùng combined.csv,
            # không lọc/sắp xếp lại — xem pillars.composite_x_marriage_age).
            mg_ref = data_refs["marriage_age_bracket"]
            ws.cell(row=row, column=1, value="Theo nhóm tuổi kết hôn (Q6)").font = SECTION_FONT
            row += 1
            for code, label in (("<18", "Kết hôn trước 18 tuổi (tảo hôn)"), (">=18", "Kết hôn từ 18 tuổi")):
                count_formula = f'COUNTIF({mg_ref.range},"={code}")'
                avg_formula = (
                    f'=IF({count_formula}=0,"n/a",'
                    f'AVERAGEIFS(B{first_data_row}:B{last_data_row},{mg_ref.range},"={code}"))'
                )
                ws.cell(row=row, column=1, value=label)
                cell = ws.cell(row=row, column=2, value=avg_formula)
                cell.number_format = "0.0"
                ws.cell(row=row, column=3, value=f"={count_formula}")
                row += 1
            row += 1

    for col_letter, width in [("A", 60), ("B", 16), ("C", 14)]:
        ws.column_dimensions[col_letter].width = width
    return {"Q14": r14, "Q32": r32}


def write_cover_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Trang bìa/tổng quan (26/07 tối, phản hồi khách "làm đẹp/chuyên nghiệp hơn") — sheet
    ĐẦU TIÊN khách nhìn thấy khi mở file, thay vì rơi thẳng vào sheet dữ liệu/tần suất. Số
    liệu ở đây là ảnh chụp tĩnh (COUNTA sống cho tổng mẫu, còn lại là text mô tả) — không
    thay thế các sheet phân tích, chỉ định hướng."""
    ws = wb.create_sheet("Tổng quan", 0)
    ws.sheet_view.showGridLines = False
    for col_letter, width in [("A", 4), ("B", 46), ("C", 46)]:
        ws.column_dimensions[col_letter].width = width

    band = PatternFill("solid", fgColor="1F4E78")
    for col in range(1, 4):
        for r in range(1, 4):
            ws.cell(row=r, column=col).fill = band
    title_cell = ws.cell(row=2, column=2, value="Thị trường, chuỗi giá trị, rào cản sản xuất và")
    title_cell.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws.cell(row=3, column=2, value="môi trường chính sách đối với cây dược liệu").font = Font(name="Arial", bold=True, size=16, color="FFFFFF")

    row = 5
    ws.cell(row=row, column=2, value="Khảo sát 85 phiếu bán cấu trúc với phụ nữ trồng/kinh doanh cây dược liệu — Lào Cai và Lai Châu").font = Font(name="Arial", italic=True, size=11, color="444444")
    row += 2

    ws.cell(row=row, column=2, value="N tổng mẫu").font = Font(name="Arial", bold=True)
    ws.cell(row=row, column=3, value=f"=COUNTA('Dữ liệu (ẩn danh)'!$A$2:$A${1+len(df)})").font = Font(name="Arial")
    row += 1
    ws.cell(row=row, column=2, value="Tỉnh khảo sát").font = Font(name="Arial", bold=True)
    ws.cell(row=row, column=3, value="Lào Cai, Lai Châu").font = Font(name="Arial")
    row += 1
    ws.cell(row=row, column=2, value="Đơn vị phân tích").font = Font(name="Arial", bold=True)
    ws.cell(row=row, column=3, value="Hộ/phụ nữ trồng, thu hái, chế biến, kinh doanh cây dược liệu").font = Font(name="Arial")
    row += 2

    ws.cell(row=row, column=2, value="NỘI DUNG FILE").font = Font(name="Arial", bold=True, size=12, color="1F4E78")
    row += 1
    contents = [
        ("A. Thị trường", "Nguồn thu nhập, mức độ phụ thuộc dược liệu, kinh nghiệm × thu nhập"),
        ("B. Chuỗi giá trị", "Khâu tham gia, ai làm/ai quyết, học thức/lãnh đạo/đi lại/thiết bị số"),
        ("C. Rào cản", "Rào cản sản xuất theo tỉnh, theo khâu tham gia, theo dân tộc"),
        ("D. Chính sách", "Vốn vay, tập huấn, hội đoàn thể, SWOT, theo dân tộc"),
        ("E. Chỉ số vai trò", "Cronbach's alpha, chỉ số tổng hợp, theo nhóm tuổi kết hôn"),
        ("Biểu đồ trụ cột / Biểu đồ", "Biểu đồ Excel gốc cho các bảng chính"),
        ("Thống kê trải phẳng", "Tần suất/% cho TOÀN BỘ ~99 biến — lớp nền/tra cứu"),
        ("Dữ liệu đã số hóa", "Bản đầy đủ có thông tin định danh (PII)"),
    ]
    for name, desc in contents:
        ws.cell(row=row, column=2, value=name).font = Font(name="Arial", bold=True, size=10, color="1F4E78")
        ws.cell(row=row, column=3, value=desc).font = Font(name="Arial", size=10)
        row += 1
    row += 1
    ws.cell(
        row=row, column=2,
        value="Mọi số liệu ở các sheet phân tích là công thức Excel sống, tham chiếu sheet "
              "\"Dữ liệu (ẩn danh)\" — sửa 1 ô dữ liệu, số liệu tự cập nhật. Mẫu 85 phiếu mang "
              "tính gợi ý cho vùng khảo sát, không đại diện thống kê cho toàn vùng trồng dược liệu.",
    ).font = Font(name="Arial", italic=True, size=9, color="666666")
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 45
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)


def write_pillar_charts_sheet(wb: Workbook, pillar_locations: dict[str, FreqBlockLocation], specs: list[dict]) -> None:
    ws = wb.create_sheet("Biểu đồ trụ cột")
    anchor_row = 1
    for spec in specs:
        col = spec["column"]
        loc = pillar_locations.get(col)
        if loc is None:
            continue
        chart = BarChart()
        chart.type = "bar"
        # 26/07 (tối): "column" có thể là 1 qid gộp (Q7/Q28/Q22a — không có trong CODEBOOK,
        # chỉ các cột đã nổ Q7_<opt> mới có) — cho phép spec khai báo "title" tường minh,
        # rơi về tra CODEBOOK như cũ cho cột đơn (age_bracket, Q8, Q30, Q21a...).
        chart.title = spec.get("title") or CODEBOOK[col]["label"]
        chart.height = 7
        chart.width = 15
        cats = Reference(wb[loc.sheet], min_col=1, min_row=loc.first_data_row, max_row=loc.last_data_row)
        data = Reference(wb[loc.sheet], min_col=2, min_row=loc.header_row, max_row=loc.last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        ws.add_chart(chart, f"A{anchor_row}")
        anchor_row += 16
