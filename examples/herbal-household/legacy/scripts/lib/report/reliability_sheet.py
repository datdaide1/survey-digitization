"""Sheet "Chỉ số tổng hợp" (Tầng 6, §7) — Cronbach's alpha cho Q14/Q32 + chỉ số tổng
hợp (nếu alpha đạt ngưỡng 0.7) kèm mô tả + so sánh nhanh theo tỉnh.

26/07 (phản hồi khách, quyết định: chuyển phần đơn giản sang công thức sống): Cronbach's
alpha (chọn CÓ tạo chỉ số hay không) vẫn TĨNH — công thức phương sai lồng nhau cho việc
"chọn ngưỡng" không đáng để làm sống. NHƯNG cột "Chỉ số (%)" của TỪNG phiếu (khi đã đạt
ngưỡng) giờ là công thức Excel sống thật — mỗi phiếu chỉ là % số dòng "vợ"/"cả hai" (hay
"vợ"/"cùng quyết định") trên tổng số dòng, tương đương AVERAGE(các cờ 0/1 mỗi dòng) — quá
đơn giản để không tận dụng công thức sống, khác hẳn Cramér's V/factor analysis/phân cụm ở
Tầng 3/4/5/7 (xem lý do kỹ thuật trong association.py). Sửa 1 ô Q14_.../Q32_... ở sheet
"Dữ liệu (ẩn danh)" là chỉ số (%) tương ứng tự cập nhật ngay, không cần chạy lại script.
"""

from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill

from .reliability import Q14_POSITIVE_CODES, Q14_ROWS, Q32_POSITIVE_CODES, Q32_ROWS, analyze
from .xlsx_writer import DataRef

TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SECTION_FONT = Font(bold=True, size=11)
NOTE_FONT = Font(italic=True, size=9, color="666666")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _item_positive_formula(cell_addr: str, positive_codes: set[str]) -> str:
    """1 nếu ô ghi đúng 1 trong các code "dương" (vd 'vo'/'ca_hai'), kể cả khi ô là tổ hợp
    multi-mark 'vo+chong' — bọc "+" trước/sau cả 2 vế để SEARCH khớp đúng ranh giới code,
    không dính substring giả (cùng nguyên tắc _combo_criteria_literals ở xlsx_writer.py,
    viết lại dạng 1 công thức IF/SEARCH thay vì nhiều COUNTIF vì đây là test từng Ô, không
    phải đếm cả 1 CỘT). Ô trống -> "++" -> không khớp code nào -> 0, giống Python
    (reliability.binary_item_matrix fillna(0.0)).
    """
    checks = [f'ISNUMBER(SEARCH("+{code}+","+"&{cell_addr}&"+"))' for code in sorted(positive_codes)]
    return f'IF(OR({",".join(checks)}),1,0)'


def _composite_formula(data_refs: dict[str, DataRef], qid: str, rows: list[str], positive_codes: set[str], record_row: int) -> str:
    terms = []
    for row_code in rows:
        ref = data_refs[f"{qid}_{row_code}"]
        cell_addr = f"'{ref.sheet}'!${ref.col_letter}${record_row}"
        terms.append(_item_positive_formula(cell_addr, positive_codes))
    return f'=AVERAGE({",".join(terms)})*100'


def write_reliability_sheet(wb: Workbook, df: pd.DataFrame, data_refs: dict[str, DataRef]) -> dict[str, Any]:
    r14 = analyze(df, "Q14", "Q14 – Chỉ số phân công lao động (18 việc, 'vợ' hoặc 'cả hai')", Q14_ROWS, Q14_POSITIVE_CODES)
    r32 = analyze(df, "Q32", "Q32 – Chỉ số tham gia quyết định của phụ nữ (8 vấn đề, 'vợ' hoặc 'cùng quyết định')", Q32_ROWS, Q32_POSITIVE_CODES)
    specs = {"Q14": (Q14_ROWS, Q14_POSITIVE_CODES), "Q32": (Q32_ROWS, Q32_POSITIVE_CODES)}

    ws = wb.create_sheet("Chỉ số tổng hợp")
    ws.cell(row=1, column=1, value="TẦNG 6 — CHỈ SỐ TỔNG HỢP (CRONBACH'S ALPHA)").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="Kiểm tra xem nhiều dòng trong 1 câu ma trận có đủ nhất quán để gộp thành 1 "
              "chỉ số duy nhất không, thay vì đọc từng dòng riêng. Mốc tham khảo: alpha >= 0.7 "
              "(Nunnally) — đạt thì tạo chỉ số tổng hợp, không đạt thì giữ nguyên từng dòng.",
    ).font = NOTE_FONT
    ws.cell(
        row=3, column=1,
        value="Cronbach's alpha (tính CÓ tạo chỉ số hay không) là số tĩnh. Cột 'Chỉ số (%)' "
              "từng phiếu bên dưới LÀ CÔNG THỨC EXCEL SỐNG — sửa dữ liệu gốc là số tự đổi.",
    ).font = NOTE_FONT
    row = 5

    for qid, r in (("Q14", r14), ("Q32", r32)):
        ws.cell(row=row, column=1, value=r.label).font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="Số dòng (item)")
        ws.cell(row=row, column=2, value=r.n_items)
        row += 1
        ws.cell(row=row, column=1, value="Cronbach's alpha")
        ws.cell(row=row, column=2, value=round(r.alpha, 3) if r.alpha is not None else "Không tính được")
        row += 1
        ws.cell(row=row, column=1, value="Quyết định")
        ws.cell(row=row, column=2, value="ĐẠT ngưỡng — tạo chỉ số tổng hợp" if r.composite is not None else "CHƯA đạt ngưỡng — giữ nguyên từng dòng")
        row += 2

        if r.composite is not None:
            # 26/07: bỏ so sánh trung bình theo tỉnh (Lào Cai/Lai Châu) — quyết định của
            # user, không cần so sánh giữa các vùng nữa. Xem
            # [[project-survey-no-region-comparison]]. Giữ lại cột "Tỉnh" ở bảng chi tiết
            # từng phiếu (chỉ để tham khảo/tra cứu, không phải so sánh).
            ws.cell(row=row, column=1, value="record_id").font = HEADER_FONT
            ws.cell(row=row, column=1).fill = HEADER_FILL
            ws.cell(row=row, column=2, value="Chỉ số (%)").font = HEADER_FONT
            ws.cell(row=row, column=2).fill = HEADER_FILL
            ws.cell(row=row, column=3, value="Tỉnh").font = HEADER_FONT
            ws.cell(row=row, column=3).fill = HEADER_FILL
            first_data_row = row + 1
            record_ids = df["record_id"]
            provinces = df["province"]
            rows_def, positive_codes = specs[qid]
            # Vị trí dòng của phiếu idx trong sheet "Dữ liệu (ẩn danh)" = data_refs bất kỳ
            # cột nào của qid đều dùng chung first_row/last_row (cùng 1 df) — lấy mốc từ
            # cột đầu tiên của rows_def để tính record_row theo đúng thứ tự df.
            base_ref = data_refs[f"{qid}_{rows_def[0]}"]
            for pos, idx in enumerate(r.composite.index):
                row += 1
                record_row = base_ref.first_row + pos
                ws.cell(row=row, column=1, value=record_ids.loc[idx])
                cell = ws.cell(row=row, column=2, value=_composite_formula(data_refs, qid, rows_def, positive_codes, record_row))
                cell.number_format = "0.0"
                ws.cell(row=row, column=3, value=provinces.loc[idx])
            last_data_row = row
            row += 1

            desc_row = row
            ws.cell(row=row, column=1, value="Trung bình toàn thể")
            avg_cell = ws.cell(row=row, column=2, value=f"=AVERAGE(B{first_data_row}:B{last_data_row})")
            avg_cell.number_format = "0.0"
            row += 2

    for col_letter, width in [("A", 34), ("B", 16), ("C", 14)]:
        ws.column_dimensions[col_letter].width = width

    return {"Q14": r14, "Q32": r32}
