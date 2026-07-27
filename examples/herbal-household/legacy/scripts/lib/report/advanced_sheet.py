"""Sheet "Phân tích nâng cao" (Tầng 7, §8 docs/implement-plan-statistics-and-client-report.md)
— factor analysis (Q14, Q32) + cluster analysis (85 phiếu). Giá trị TĨNH, tính 1 lần
bằng script Python (eigen-decomposition, K-means) — không có công thức Excel gốc cho 2
kỹ thuật này (§8.1), khác các sheet Tầng 3-6 vốn sống bằng công thức. Ảnh (scree plot,
scatter) chèn như ảnh thường (Insert Picture), bảng dán giá trị thường — khách copy/
chụp màn hình được như mọi bảng/ảnh khác trong file.
"""

from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from . import mpl_charts as charts
from .cluster_analysis import run_cluster_analysis
from .codebook import build_codebook
from .factor_analysis import FactorAnalysisResult, run_factor_analysis
from .reliability import Q14_POSITIVE_CODES, Q14_ROWS, Q32_POSITIVE_CODES, Q32_ROWS

CODEBOOK = build_codebook()
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SECTION_FONT = Font(bold=True, size=11)
NOTE_FONT = Font(italic=True, size=9, color="666666")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")


def _item_label(qid: str, row_code: str) -> str:
    return CODEBOOK[f"{qid}_{row_code}"]["label"].split(" — ", 1)[-1]


def _write_factor_block(ws: Worksheet, row: int, result: FactorAnalysisResult, rows: list[str], tmp: Path) -> int:
    ws.cell(row=row, column=1, value=result.label).font = SECTION_FONT
    row += 1
    ws.cell(
        row=row, column=1,
        value=f"Số câu (item): {result.n_items} | Số phiếu: {result.n_obs} | "
              f"Số nhân tố giữ lại (eigenvalue > 1): {result.n_factors}",
    )
    row += 1
    if result.low_ratio_warning:
        c = ws.cell(
            row=row, column=1,
            value="Tỷ lệ số câu/số phiếu khá mỏng (< 5) — kết quả mục này chỉ mang tính "
                  "minh hoạ, chưa đủ tin cậy để kết luận chắc chắn.",
        )
        c.font = NOTE_FONT
        c.fill = WARNING_FILL
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Câu hỏi (item)").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for j, col_name in enumerate(result.loadings.columns):
        c = ws.cell(row=row, column=2 + j, value=col_name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    for row_code in rows:
        ws.cell(row=row, column=1, value=_item_label(result.qid, row_code))
        for j, col_name in enumerate(result.loadings.columns):
            ws.cell(row=row, column=2 + j, value=round(float(result.loadings.loc[row_code, col_name]), 2))
        row += 1
    row += 1

    png = charts.scree_plot(result.eigenvalues, tmp / f"scree_{result.qid}.png", f"Scree plot — {result.label}")
    img = XLImage(png)
    img.width, img.height = 460, 260
    ws.add_image(img, f"A{row}")
    row += 15
    return row


def write_advanced_sheet(wb: Workbook, df: pd.DataFrame) -> dict[str, Any]:
    tmp = Path(tempfile.mkdtemp(prefix="xlsx_advanced_"))
    ws = wb.create_sheet("Phân tích nâng cao")

    ws.cell(row=1, column=1, value="TẦNG 7 — PHÂN TÍCH NÂNG CAO").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="Tính 1 lần bằng script Python (eigen-decomposition, K-means) — không sống "
              "theo công thức Excel như các tầng khác, cần chạy lại script khi có dữ liệu "
              "mới. Cỡ mẫu 85 phiếu là nhỏ cho 2 kỹ thuật này — kết quả mang tính gợi ý/"
              "thăm dò, KHÔNG phải phân loại chính thức.",
    ).font = NOTE_FONT
    row = 4

    ws.cell(row=row, column=1, value="FACTOR ANALYSIS").font = TITLE_FONT
    row += 2
    r14 = run_factor_analysis(df, "Q14", "Q14 – Phân công lao động (18 việc)", Q14_ROWS, Q14_POSITIVE_CODES)
    row = _write_factor_block(ws, row, r14, Q14_ROWS, tmp)
    r32 = run_factor_analysis(df, "Q32", "Q32 – Ra quyết định (8 vấn đề)", Q32_ROWS, Q32_POSITIVE_CODES)
    row = _write_factor_block(ws, row, r32, Q32_ROWS, tmp)

    ws.cell(row=row, column=1, value="CLUSTER ANALYSIS").font = TITLE_FONT
    row += 2
    cluster = run_cluster_analysis(df)

    ws.cell(row=row, column=1, value="Số cụm (k) đã thử và điểm silhouette (càng gần 1 càng rõ cụm):").font = SECTION_FONT
    row += 1
    for k in cluster.k_tried:
        ws.cell(row=row, column=1, value=f"k = {k}")
        ws.cell(row=row, column=2, value=round(cluster.silhouette_scores[k], 3))
        if k == cluster.best_k:
            ws.cell(row=row, column=3, value="Được chọn (silhouette cao nhất)")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value=f"Đặc điểm từng cụm (k = {cluster.best_k}, giá trị trung bình, đơn vị gốc):").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Đặc trưng").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    cluster_ids = list(cluster.cluster_means_raw.index)
    for j, cid in enumerate(cluster_ids):
        c = ws.cell(row=row, column=2 + j, value=f"Cụm {cid + 1} (n={cluster.cluster_sizes[cid]})")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    for feat in cluster.feature_labels:
        ws.cell(row=row, column=1, value=feat)
        for j, cid in enumerate(cluster_ids):
            ws.cell(row=row, column=2 + j, value=round(float(cluster.cluster_means_raw.loc[cid, feat]), 2))
        row += 1
    row += 1

    png = charts.cluster_scatter(cluster.pca_2d, cluster.labels, tmp / "cluster_scatter.png", "Phân cụm 85 phiếu (giảm chiều 2D)")
    img = XLImage(png)
    img.width, img.height = 460, 360
    ws.add_image(img, f"A{row}")

    ws.column_dimensions["A"].width = 34
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 20

    return {"factor": {"Q14": r14, "Q32": r32}, "cluster": cluster}
