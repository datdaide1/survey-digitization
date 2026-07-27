"""Sheet "Effect size & so sánh nhóm" (Tầng 4 §5 + Tầng 5 §6) — giá trị TĨNH (cùng lý
do kỹ thuật với Tầng 3, xem scripts/lib/report/association.py) + forest plot cho các
cặp odds ratio (native Excel chart, khách copy/phóng to được).

CẢNH BÁO bắt buộc theo §5/§6: đây là "liên hệ quan sát được", không phải bằng chứng
nhân quả; subgroup nhỏ (n<10) chỉ mang tính tham khảo — áp dụng nhất quán ở cả sheet
này lẫn DOCX (xem docx_writer._add_effect_size_section).
"""

from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from . import mpl_charts as charts
from .codebook import build_codebook
from .curated_pairs import EFFECT_SIZE_PAIRS, GROUP_COMPARISON_PAIRS, derive_helper_columns, resolve_label
from .effect_size import compute_effect_size
from .nonparametric import kruskal_wallis, mann_whitney

CODEBOOK = build_codebook()
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SECTION_FONT = Font(bold=True, size=11)
NOTE_FONT = Font(italic=True, size=9, color="666666")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LOW_N_FILL = PatternFill("solid", fgColor="D9D9D9")


def _label(col: str) -> str:
    # 26/07 (phản hồi khách): trước đây rơi thẳng về tên cột thô (vd "Q11_la_hoi_vien")
    # cho các cột phụ trợ derive_helper_columns() tạo ra — không có trong CODEBOOK gốc.
    # Tra qua resolve_label() (CODEBOOK trước, HELPER_COLUMN_LABELS sau) để luôn ra tiếng
    # Việt đọc được.
    return resolve_label(col, CODEBOOK)


def write_effect_size_sheet(wb: Workbook, df: pd.DataFrame) -> dict[str, Any]:
    df2 = derive_helper_columns(df)
    ws = wb.create_sheet("Effect size & so sánh nhóm")

    ws.cell(row=1, column=1, value="TẦNG 4 — ĐỘ ẢNH HƯỞNG (EFFECT SIZE), CÓ CHIỀU").font = TITLE_FONT
    ws.cell(
        row=2, column=1,
        value="KHÔNG PHẢI bằng chứng nhân quả — khảo sát cắt ngang, chỉ nói lên liên hệ "
              "quan sát được, đã tính riêng cho từng cặp (không kiểm soát yếu tố khác).",
    ).font = NOTE_FONT
    ws.cell(
        row=3, column=1,
        value="LƯU Ý: số trong sheet này (cả Tầng 4 lẫn Tầng 5 bên dưới) là TÍNH SẴN bằng "
              "Python (odds ratio/eta-squared/Mann-Whitney/Kruskal-Wallis), KHÔNG phải công "
              "thức Excel sống — sửa dữ liệu gốc không tự cập nhật, phải chạy lại "
              "scripts/build_client_report.py. Khác sheet 'Thống kê tổng hợp' (Tầng 1) và "
              "biểu đồ Phần A, vốn sống bằng công thức Excel.",
    ).font = NOTE_FONT
    row = 5

    header_row = row
    headers = ["Chủ đề", "Biến 1", "Biến 2", "Loại", "Giá trị", "n", "Diễn giải gốc (§3.2)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=header_row, column=1 + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1

    effect_rows = []
    odds_ratio_rows_for_chart = []
    for topic, a, b, kind, desc in EFFECT_SIZE_PAIRS:
        r = compute_effect_size(kind, a, b, df2)
        ws.cell(row=row, column=1, value=topic)
        ws.cell(row=row, column=2, value=_label(a))
        ws.cell(row=row, column=3, value=_label(b))
        ws.cell(row=row, column=4, value=kind)
        value_cell = ws.cell(row=row, column=5, value=round(r.value, 3) if r.value is not None else None)
        ws.cell(row=row, column=6, value=r.n)
        ws.cell(row=row, column=7, value=desc)
        if kind == "odds_ratio" and r.value is not None:
            odds_ratio_rows_for_chart.append((row, f"{_label(a)} → {_label(b)}"))
        effect_rows.append({"topic": topic, "a": a, "b": b, "kind": kind, "result": r, "desc": desc})
        row += 1

    or_start = odds_ratio_rows_for_chart[0][0] if odds_ratio_rows_for_chart else None
    or_end = odds_ratio_rows_for_chart[-1][0] if odds_ratio_rows_for_chart else None
    if or_start:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Forest plot — Odds ratio (mốc trung lập = 1)"
        chart.height = 8
        chart.width = 16
        data = Reference(ws, min_col=5, min_row=header_row, max_row=or_end)
        cats = Reference(ws, min_col=2, min_row=or_start, max_row=or_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"I{header_row}")

    row += 2
    ws.cell(row=row, column=1, value="TẦNG 5 — SO SÁNH NHÓM (KIỂM ĐỊNH PHI THAM SỐ)").font = TITLE_FONT
    row += 1
    ws.cell(
        row=row, column=1,
        value="Mann-Whitney U (2 nhóm) / Kruskal-Wallis (3+ nhóm) — không giả định phân phối "
              "chuẩn. Nhóm n<10 chỉ mang tính tham khảo, không diễn giải 'khác biệt có ý nghĩa "
              "thống kê' theo nghĩa hàn lâm.",
    ).font = NOTE_FONT
    row += 1
    ws.cell(
        row=row, column=1,
        value="Cách chọn: các cặp bên dưới là VÍ DỤ MINH HOẠ chọn sẵn khi lên kế hoạch báo "
              "cáo, KHÔNG phải kết quả quét tự động toàn bộ dữ liệu để tìm khác biệt lớn "
              "nhất — cần so sánh thêm nhóm nào khác thì yêu cầu bổ sung cụ thể.",
    ).font = NOTE_FONT
    row += 2

    tmp = Path(tempfile.mkdtemp(prefix="xlsx_boxwhisker_"))
    group_results = []
    for desc, numcol, groupcol, kind in GROUP_COMPARISON_PAIRS:
        fn = mann_whitney if kind == "mann_whitney" else kruskal_wallis
        r = fn(df2[numcol], df2[groupcol])
        ws.cell(row=row, column=1, value=desc).font = SECTION_FONT
        row += 1
        header_row2 = row
        for j, h in enumerate(["Nhóm", "n", "Trung vị"]):
            c = ws.cell(row=row, column=1 + j, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        row += 1
        for g in r.groups:
            ws.cell(row=row, column=1, value=str(g["group"]))
            ws.cell(row=row, column=2, value=g["n"])
            ws.cell(row=row, column=3, value=round(g["median"], 2) if g["median"] is not None else None)
            if g["n"] < 10:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = LOW_N_FILL
            row += 1
        ws.cell(row=row, column=1, value=f"Thống kê kiểm định ({r.test}) = {r.statistic:.3f}" if r.statistic is not None else "Không đủ dữ liệu")
        ws.cell(row=row, column=3, value=f"p = {r.p_value:.4f}" if r.p_value is not None else None)
        row += 2

        if any(g.get("values") for g in r.groups):
            groups_for_plot = [{"label": str(g["group"]), "values": g["values"]} for g in r.groups if g.get("values")]
            png = charts.box_whisker(desc, groups_for_plot, tmp / f"box_{numcol}_{groupcol}.png")
            img = XLImage(png)
            target_width = 420
            img.height = int(target_width * img.height / img.width)
            img.width = target_width
            ws.add_image(img, f"A{row}")
            row += 14

        group_results.append({"desc": desc, "result": r})

    for col_letter, width in [("A", 26), ("B", 30), ("C", 30), ("D", 14), ("E", 12), ("F", 8), ("G", 60)]:
        ws.column_dimensions[col_letter].width = width

    return {"effect_rows": effect_rows, "group_results": group_results}
