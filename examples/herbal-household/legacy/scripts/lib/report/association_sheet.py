"""Sheet "Ma trận liên quan" (Tầng 3, §4 + §4.2) — heatmap từng khối A/B1/B2/B3, tóm
tắt liên khối 4×4, bảng Top liên quan mạnh nhất, và ma trận đầy đủ ~99×99. Giá trị TĨNH
(tính bằng scripts/lib/report/association.py, xem docstring ở đó để biết lý do không
dùng công thức Excel sống cho tầng này) — tô màu bằng conditional formatting color
scale thật của Excel, vẫn là heatmap đúng nghĩa dù số không tự cập nhật.
"""

from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .association import (
    LOW_N_THRESHOLD,
    AssociationResult,
    compute_association_matrix,
    top_associations,
)
from .codebook import build_codebook

CODEBOOK = build_codebook()

TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SECTION_FONT = Font(bold=True, size=11)
NOTE_FONT = Font(italic=True, size=9, color="666666")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LOW_N_FILL = PatternFill("solid", fgColor="D9D9D9")

SECTION_TITLES = {"A": "Khối A — Thông tin chung", "B1": "Khối B1 — Phân công lao động", "B2": "Khối B2 — Tiếp cận nguồn lực", "B3": "Khối B3 — Ra quyết định"}

METHODOLOGY_NOTE = (
    "Chỉ số tương quan/liên hệ mang tính MÔ TẢ/THĂM DÒ, không phải bằng chứng quan hệ "
    "nhân quả — vẫn có kèm p-value (kiểm định thống kê) ở bảng 'Top liên quan mạnh nhất' "
    "bên dưới để biết liên hệ quan sát được có dễ xảy ra do ngẫu nhiên hay không. Cramér's "
    "V (0 đến 1, biến phân loại-phân loại), Spearman (-1 đến 1, biến số-biến số), "
    "rank-biserial (-1 đến 1, biến số-nhị phân), eta (0 đến 1, biến số-phân loại nhiều "
    "mức). Ô có n<20 được tô xám — độ tin cậy thấp, p-value (nếu có) không đáng tin."
)

# 26/07 (phản hồi khách: "top liên quan toàn p=0,000 mà chả liên quan gì, kiểu tuổi với sở
# hữu thiết bị-chồng"): Cramér's V/rank-biserial có thể vọt lên gần 1 (kèm p≈0) một cách
# GIẢ TẠO nếu 1 trong 2 biến có 1 nhóm cực nhỏ (vd biến chỉ có 2/85 phiếu ở 1 mức) — chỉ
# cần 2 phiếu đó tình cờ trùng nhóm nào đó ở biến kia là ra ngay. Giờ LOẠI KHỎI bảng "Top
# liên quan mạnh nhất" mọi cặp có nhóm nhỏ nhất <5 phiếu ở 1 trong 2 biến (không chỉ lọc
# theo n chung <20 như trước) — xem MIN_CATEGORY_N trong association.py. Bảng dưới đây vì
# vậy sẽ có ít cặp hơn bản trước, nhưng đáng tin hơn.
RELIABILITY_NOTE = (
    "Bảng 'Top liên quan mạnh nhất' đã LOẠI các cặp mà 1 trong 2 biến có nhóm nhỏ nhất "
    "<5 phiếu (vd biến chỉ 2/85 phiếu thuộc 1 mức) — nhóm quá nhỏ dễ cho V=1/p=0,000 giả "
    "tạo, không phản ánh liên hệ thật. Ma trận đầy đủ ~99×99 bên dưới vẫn giữ mọi cặp "
    "(kể cả nhóm nhỏ) để tham khảo, chỉ tô xám các ô không đáng tin."
)

# 26/07 (phản hồi khách): giá trị tĩnh trong sheet này KHÔNG tự cập nhật khi sửa dữ liệu ở
# sheet "Dữ liệu (ẩn danh)" — khác Tầng 1/biểu đồ Phần A (công thức Excel sống). Lý do kỹ
# thuật: xem docstring đầu scripts/lib/report/association.py. Ghi rõ trong sheet để khách
# không hiểu nhầm là công thức sống.
STATIC_VALUE_NOTE = (
    "LƯU Ý: các số trong sheet này được TÍNH SẴN bằng Python khi tạo báo cáo (Cramér's V/"
    "eta/rank-biserial cần hàng chục nghìn công thức Excel lồng nhau mới làm 'sống' được "
    "cho ~99×99 cặp biến — quá rủi ro sai sót nên không làm vậy). Sửa dữ liệu gốc ở sheet "
    "'Dữ liệu (ẩn danh)' sẽ KHÔNG tự cập nhật số ở đây — phải chạy lại "
    "scripts/build_client_report.py để có bản mới. Ngược lại, sheet 'Thống kê tổng hợp' và "
    "biểu đồ Phần A dùng công thức Excel sống, sửa dữ liệu là số tự đổi ngay."
)


def _value_matrix(columns: list[str], matrix: dict[tuple[str, str], AssociationResult]) -> pd.DataFrame:
    n = len(columns)
    data = [[None] * n for _ in range(n)]
    for i, a in enumerate(columns):
        for j, b in enumerate(columns):
            if i == j:
                continue
            key = (a, b) if (a, b) in matrix else (b, a)
            r = matrix.get(key)
            if r is not None:
                data[i][j] = r.value
    return pd.DataFrame(data, index=columns, columns=columns)


def _write_indexed_heatmap(
    ws: Worksheet, start_row: int, start_col: int, columns: list[str],
    matrix: dict[tuple[str, str], AssociationResult],
) -> int:
    """Viết 1 heatmap n×n với header là số thứ tự (1..n) — chú giải nhãn đầy đủ ở bên
    phải (label dài tiếng Việt không đọc nổi nếu làm header cột cho ma trận lớn)."""
    n = len(columns)
    value_df = _value_matrix(columns, matrix)

    # Header hàng/cột bằng số thứ tự
    for j in range(n):
        c = ws.cell(row=start_row, column=start_col + 1 + j, value=j + 1)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(text_rotation=90, horizontal="center")
    for i, col in enumerate(columns):
        c = ws.cell(row=start_row + 1 + i, column=start_col, value=i + 1)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    for i in range(n):
        for j in range(n):
            cell = ws.cell(row=start_row + 1 + i, column=start_col + 1 + j)
            if i == j:
                cell.value = 1.0
                continue
            val = value_df.iat[i, j]
            if val is None:
                continue
            cell.value = round(float(val), 3)
            key = (columns[i], columns[j])
            r = matrix.get(key) or matrix.get((columns[j], columns[i]))
            if r is not None and r.low_n:
                cell.fill = LOW_N_FILL

    data_range = (
        f"{get_column_letter(start_col + 1)}{start_row + 1}:"
        f"{get_column_letter(start_col + n)}{start_row + n}"
    )
    rule = ColorScaleRule(
        start_type="num", start_value=-1, start_color="F8696B",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=1, end_color="5A8AC6",
    )
    ws.conditional_formatting.add(data_range, rule)

    # Chú giải số thứ tự -> nhãn đầy đủ, đặt bên phải heatmap
    legend_col = start_col + n + 2
    ws.cell(row=start_row, column=legend_col, value="STT").font = HEADER_FONT
    ws.cell(row=start_row, column=legend_col + 1, value="Biến").font = HEADER_FONT
    for i, col in enumerate(columns):
        ws.cell(row=start_row + 1 + i, column=legend_col, value=i + 1)
        ws.cell(row=start_row + 1 + i, column=legend_col + 1, value=f"{CODEBOOK[col]['label']} [{col}]")

    return start_row + n + 3


def write_association_sheet(wb: Workbook, df: pd.DataFrame, all_columns: list[str]) -> dict[str, Any]:
    ws = wb.create_sheet("Ma trận liên quan")
    ws.cell(row=1, column=1, value="TẦNG 3 — MA TRẬN LIÊN QUAN GIỮA CÁC BIẾN").font = TITLE_FONT
    ws.cell(row=2, column=1, value=METHODOLOGY_NOTE).font = NOTE_FONT
    ws.cell(row=3, column=1, value=STATIC_VALUE_NOTE).font = NOTE_FONT
    row = 5

    full_matrix = compute_association_matrix(df, all_columns)

    section_columns: dict[str, list[str]] = {"A": [], "B1": [], "B2": [], "B3": []}
    for col in all_columns:
        section = CODEBOOK[col].get("section")
        if section in section_columns:
            section_columns[section].append(col)

    for section, cols in section_columns.items():
        ws.cell(row=row, column=1, value=SECTION_TITLES[section]).font = SECTION_FONT
        row += 1
        row = _write_indexed_heatmap(ws, row, 1, cols, full_matrix)

    # --- Tóm tắt liên khối 4x4 ---
    ws.cell(row=row, column=1, value="Tóm tắt liên khối (điểm liên quan trung bình |giá trị| giữa 2 khối)").font = SECTION_FONT
    row += 1
    block_names = list(section_columns.keys())
    header_row = row
    for j, b in enumerate(block_names):
        ws.cell(row=header_row, column=2 + j, value=b).font = HEADER_FONT
        ws.cell(row=header_row, column=2 + j).fill = HEADER_FILL
    for i, b1 in enumerate(block_names):
        ws.cell(row=header_row + 1 + i, column=1, value=b1).font = HEADER_FONT
        ws.cell(row=header_row + 1 + i, column=1).fill = HEADER_FILL
        for j, b2 in enumerate(block_names):
            vals = []
            for a in section_columns[b1]:
                for b in section_columns[b2]:
                    if a == b:
                        continue
                    r = full_matrix.get((a, b)) or full_matrix.get((b, a))
                    if r is not None and r.value is not None:
                        vals.append(abs(r.value))
            avg = sum(vals) / len(vals) if vals else None
            cell = ws.cell(row=header_row + 1 + i, column=2 + j, value=round(avg, 3) if avg is not None else None)
    summary_range = f"B{header_row + 1}:{get_column_letter(1 + len(block_names))}{header_row + len(block_names)}"
    ws.conditional_formatting.add(
        summary_range,
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF", end_type="num", end_value=1, end_color="5A8AC6"),
    )
    row = header_row + len(block_names) + 2

    # --- Top liên quan mạnh nhất ---
    ws.cell(row=row, column=1, value="Top 20 liên quan mạnh nhất toàn bộ dữ liệu (đã loại cặp cùng 1 câu hỏi gốc + cặp có nhóm quá nhỏ)").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value=RELIABILITY_NOTE).font = NOTE_FONT
    row += 1
    header_row = row
    headers = ["Biến 1", "Biến 2", "Phương pháp", "Giá trị", "n", "p-value", "Ý nghĩa thống kê (p<0,05)", "Ghi chú"]
    for j, h in enumerate(headers):
        ws.cell(row=header_row, column=1 + j, value=h).font = HEADER_FONT
        ws.cell(row=header_row, column=1 + j).fill = HEADER_FILL
    row += 1
    top20 = top_associations(full_matrix, 20)
    for a, b, r in top20:
        ws.cell(row=row, column=1, value=CODEBOOK[a]["label"])
        ws.cell(row=row, column=2, value=CODEBOOK[b]["label"])
        ws.cell(row=row, column=3, value=r.method)
        ws.cell(row=row, column=4, value=round(r.value, 3))
        ws.cell(row=row, column=5, value=r.n)
        ws.cell(row=row, column=6, value=round(r.p_value, 4) if r.p_value is not None else None)
        sig = r.significant
        ws.cell(row=row, column=7, value="Có" if sig is True else ("Chưa rõ" if sig is None else "Không"))
        ws.cell(row=row, column=8, value="n<20, độ tin cậy thấp" if r.low_n else "")
        row += 1
    row += 2

    # --- Ma trận đầy đủ ~99x99 ---
    ws.cell(row=row, column=1, value="Ma trận đầy đủ toàn bộ biến (~99×99, phụ lục kỹ thuật)").font = SECTION_FONT
    row += 1
    row = _write_indexed_heatmap(ws, row, 1, all_columns, full_matrix)

    for col_letter in ("A",):
        ws.column_dimensions[col_letter].width = 6

    return {"matrix": full_matrix, "top20": top20, "section_columns": section_columns}
