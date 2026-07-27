#!/usr/bin/env python3
"""Báo cáo thống kê gửi khách — bản 4 TRỤ CỘT (26/07), thay cho 7 tầng kỹ thuật cũ.

Đọc output/combined.csv (chạy scripts/build_stats_layer.py trước) + output/full/*.json
(cho sheet "Dữ liệu đã số hóa", có PII) -> reports/bao-cao-khao-sat.xlsx + .docx.

Cấu trúc theo docs/implement-plan-statistics-and-client-report.md (bản 26/07):
A. Thị trường & mức độ gắn bó | B. Vị trí chuỗi giá trị | C. Rào cản sản xuất |
D. Môi trường chính sách (+ SWOT) | E. Chỉ số vai trò chuỗi giá trị (Cronbach's alpha) |
Biểu đồ trụ cột | Thống kê trải phẳng (nền, giữ nguyên Tầng 1 cũ) | Biểu đồ (cũ) |
Dữ liệu đã số hóa (PII) | Dữ liệu (ẩn danh) (ẩn, nguồn công thức).

BỎ so với bản cũ: ma trận tương quan 99×99, effect size/so sánh nhóm phi tham số áp toàn
bộ, factor/cluster analysis — xem lý do ở docs/implement-plan-statistics-and-client-report.md
§2. Code cũ (association*.py, effect_size*.py, nonparametric.py, factor_analysis.py,
cluster_analysis.py, advanced_sheet.py) vẫn còn nguyên trong repo, không bị xoá.

Chạy:
  python scripts/build_pillar_report.py
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import pandas as pd  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from lib.flatten import flatten_full  # noqa: E402
from lib.records import iter_full_records  # noqa: E402
from lib.report import pillars  # noqa: E402
from lib.report.pillar_docx import build_pillar_docx  # noqa: E402
from lib.report.pillar_xlsx import (  # noqa: E402
    add_derived_columns,
    write_pillar_a_sheet,
    write_pillar_b_sheet,
    write_pillar_c_sheet,
    write_pillar_d_sheet,
    write_pillar_e_sheet,
)
from lib.report.xlsx_writer import (  # noqa: E402
    write_all_charts_sheet,
    write_data_sheet,
    write_frequency_sheet,
    write_full_data_sheet,
)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# 26/07 (đêm, phản hồi khách — "bảng có gì thì tưng đó biểu đồ chứ"): KHÔNG còn danh sách
# chart chọn lọc — write_all_charts_sheet vẽ 1 chart cho MỌI bảng trong locations. Dict dưới
# đây chỉ còn tác dụng ĐẶT TÊN đẹp cho vài qid gộp (Q7/Q28/Q22a...) không có sẵn trong
# CODEBOOK (chỉ các cột đã nổ mới có) — không dùng để LỌC bảng nào được vẽ nữa.
PILLAR_CHART_TITLES = {
    "Q7": "Q7 – Nguồn thu nhập chính",
    "Q28": "Q28 – Khó khăn khi trồng/kinh doanh dược liệu",
    "Q22a": "Q22a – Vay vốn sản xuất/kinh doanh",
}


def build_workbook(combined_csv: str, full_dir: str) -> Workbook:
    df = pd.read_csv(combined_csv)
    column_order = [c for c in df.columns if c != "record_id"]
    df_aug = add_derived_columns(df)

    full_rows = [flatten_full(r) for r in iter_full_records(full_dir)]
    full_df = pd.DataFrame(full_rows)

    wb = Workbook()
    wb.remove(wb.active)

    data_refs = write_data_sheet(wb, "Dữ liệu (ẩn danh)", df_aug, hidden=True)

    pillar_a_loc = write_pillar_a_sheet(wb, df, data_refs)
    pillar_b_q30_loc = write_pillar_b_sheet(wb, df_aug, data_refs)
    pillar_c_loc = write_pillar_c_sheet(wb, df_aug, data_refs)
    swot = pillars.pillar_d(df)["swot"]
    pillar_d_loc = write_pillar_d_sheet(wb, df, data_refs, swot)
    write_pillar_e_sheet(wb, df, data_refs)

    # 26/07 (đêm, phản hồi khách — "bảng có gì thì tưng đó biểu đồ chứ"): gộp MỌI location
    # đã theo dõi được ở A-D (kể cả Q30 riêng của B, vốn trước đây không được đưa vào chart
    # trụ cột) rồi vẽ 1 chart/bảng — không còn danh sách chọn lọc PILLAR_CHART_SPECS cũ.
    pillar_locations = {**pillar_a_loc, "Q30": pillar_b_q30_loc, **pillar_c_loc, **pillar_d_loc}
    write_all_charts_sheet(wb, "Biểu đồ trụ cột", pillar_locations, titles=PILLAR_CHART_TITLES)

    # 26/07 (phản hồi khách — "thống kê trải phẳng cho tất cả các câu đâu mất rồi"):
    # sheet này VẪN tồn tại từ bản trước (hàm write_frequency_sheet có sẵn không đổi
    # tên sheet cũ "Thống kê tổng hợp"), nhưng bị chìm ở gần cuối và tên không khớp
    # với tên đã nói trong plan/tóm tắt gửi khách ("Thống kê trải phẳng") — đổi tên cho
    # khớp + đưa lên ngay sau các sheet trụ cột A-E để dễ tìm hơn, thay vì để tít cuối
    # cạnh 2 sheet dữ liệu thô.
    # 26/07 (đêm): write_frequency_sheet giờ trả về CẢ matrix_layouts (cột n/%/lựa chọn cho
    # Q14/Q32 — bảng ma trận, cần chart nhiều chuỗi) — write_all_charts_sheet dùng nó để vẽ
    # đúng 1 chart/bảng cho MỌI khối trong "Thống kê trải phẳng", kể cả bảng ma trận.
    locations, matrix_layouts = write_frequency_sheet(wb, data_refs, df, column_order)
    # 26/07 (đêm, bug phát hiện qua ẢNH RENDER — không phải qua recalc): đổi tên sheet
    # PHẢI xảy ra TRƯỚC khi vẽ chart tham chiếu tới sheet đó. Chart series/category ref
    # (vd "'Thống kê tổng hợp'!$C$367:$C$373") lưu tên sheet dạng CHUỖI TĨNH tại thời điểm
    # tạo chart — đổi .title của worksheet SAU đó không cập nhật lại chuỗi này, khiến toàn
    # bộ chart trên sheet "Biểu đồ" tham chiếu tới 1 sheet không còn tồn tại -> chart trống
    # rỗng (không có cột, không nhãn, chỉ còn trục mặc định 0-12) dù recalc/formula vẫn
    # 0 lỗi (vì đây là lỗi tham chiếu CHART, không phải công thức Ô). Sửa: đổi tên NGAY, rồi
    # cập nhật loc.sheet trong từng FreqBlockLocation khớp tên mới, TRƯỚC khi vẽ chart.
    wb["Thống kê tổng hợp"].title = "Thống kê trải phẳng"
    for loc in locations.values():
        loc.sheet = "Thống kê trải phẳng"
    write_all_charts_sheet(wb, "Biểu đồ", locations, matrix_layouts)
    write_full_data_sheet(wb, full_df)

    # 26/07 (tối, phản hồi khách "làm đẹp/chuyên nghiệp hơn"): trang bìa/tổng quan làm sheet
    # ĐẦU TIÊN (index 0) — khách mở file thấy tổng quan có định hướng trước, không rơi
    # thẳng vào 1 sheet phân tích cụ thể. Tab color để phân biệt nhanh 4 trụ cột + chỉ số E
    # với các sheet nền/dữ liệu thô khi cuộn ngang danh sách tab.
    from lib.report.pillar_xlsx import write_cover_sheet
    write_cover_sheet(wb, df)
    tab_colors = {
        "Tổng quan": "1F4E78",
        "A. Thị trường": "4C72B0", "B. Chuỗi giá trị": "4C72B0",
        "C. Rào cản": "4C72B0", "D. Chính sách": "4C72B0", "E. Chỉ số vai trò": "4C72B0",
        "Thống kê trải phẳng": "9DB8D2",
    }
    for name, color in tab_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    wb.active = wb.sheetnames.index("Tổng quan")
    return wb


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--combined-csv", default="output/combined.csv")
    ap.add_argument("--full-dir", default="output/full")
    ap.add_argument("--xlsx-out", default="reports/bao-cao-khao-sat.xlsx")
    ap.add_argument("--docx-out", default="reports/bao-cao-khao-sat.docx")
    args = ap.parse_args()

    wb = build_workbook(args.combined_csv, args.full_dir)
    xlsx_path = Path(args.xlsx_out)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"OK: đã ghi {xlsx_path} — sheet: {', '.join(wb.sheetnames)}")

    df = pd.read_csv(args.combined_csv)
    docx_path = Path(args.docx_out)
    build_pillar_docx(df, docx_path)
    print(f"OK: đã ghi {docx_path}")
    sys.exit(0)


if __name__ == "__main__":
    main()
